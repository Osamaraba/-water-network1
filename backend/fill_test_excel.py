# -*- coding: utf-8 -*-
"""Create a test Excel with sample employees for bulk import testing."""
import os
from openpyxl import load_workbook
from datetime import datetime


def fill_test_data():
    path = "D:\\yarmouk_water_management_pro\\backend\\output\\employee_bulk_import_20260830_221646.xlsx"
    if not os.path.exists(path):
        path = [f for f in os.listdir("D:\\yarmouk_water_management_pro\\backend\\output") if f.startswith("employee_bulk_import")][0]
        path = os.path.join("D:\\yarmouk_water_management_pro\\backend\\output", path)

    wb = load_workbook(path)
    ws = wb["4-DataEntry"]

    start_row = 6

    employees = [
        # Row 6: Level 1 - General Manager (no manager)
        ["EMP101", "خالد العام", "Khaled Al-Amm", "المدير العام",
         "0791000001", "khaled@yarmouk.gov.jo", "2020-01-15",
         "general_manager", "شركة مياه اليرموك", "COMPANY", "",
         "OFFICE", "FALSE", "", "", "200", "FALSE",
         "Yarmouk@2025", "المدير العام"],

        # Row 7: Level 2 - Vice General Manager
        ["EMP102", "سارة النائبة", "Sara Al-Naiba", "نائب المدير العام",
         "0791000002", "sara@yarmouk.gov.jo", "2020-03-01",
         "general_manager", "شركة مياه اليرموك", "COMPANY", "EMP101",
         "OFFICE", "FALSE", "", "", "200", "FALSE",
         "Yarmouk@2025", "نائب المدير العام"],

        # Row 8: Level 3 - Province Director (إربد)
        ["EMP103", "أحمد إربد", "Ahmed Irbid", "مدير إدارة محافظة إربد",
         "0791000003", "ahmed@yarmouk.gov.jo", "2020-06-15",
         "hr_manager", "إدارة محافظة إربد", "PROVINCE", "EMP101",
         "OFFICE", "FALSE", "", "", "200", "FALSE",
         "Yarmouk@2025", "مدير إربد"],

        # Row 9: Level 4 - Directorate Director
        ["EMP104", "محمد الشمال", "Mohammed North", "مدير مديرية مياه الشمال",
         "0791000004", "mohammed@yarmouk.gov.jo", "2021-01-10",
         "field_supervisor", "مديرية مياه الشمال", "DIRECTORATE", "EMP103",
         "FIELD", "TRUE", "32.5550", "35.8500", "200", "FALSE",
         "Yarmouk@2025", "مدير مديرية"],

        # Row 10: Level 5 - Section Head
        ["EMP105", "فاطمة قسم 1", "Fatima Section1", "رئيس قسم 1",
         "0791000005", "fatima@yarmouk.gov.jo", "2021-05-20",
         "office_supervisor", "قسم 1 - مديرية الشمال", "SECTION", "EMP104",
         "FIELD", "TRUE", "32.5560", "35.8510", "150", "FALSE",
         "Yarmouk@2025", "رئيس قسم"],

        # Row 11: Level 6 - Subsection Chief
        ["EMP106", "علي شعبة أ", "Ali SubA", "رئيس شعبة أ",
         "0791000006", "ali@yarmouk.gov.jo", "2022-02-15",
         "office_supervisor", "شعبة أ - قسم 1", "SUBSECTION", "EMP105",
         "FIELD", "TRUE", "32.5570", "35.8520", "100", "FALSE",
         "Yarmouk@2025", "رئيس شعبة"],

        # Row 12: Level 7 - Employee
        ["EMP107", "نور موظف 1", "Noor Emp1", "موظف فني",
         "0791000007", "noor@yarmouk.gov.jo", "2023-01-15",
         "employee", "وحدة الشعبة أ", "UNIT", "EMP106",
         "FIELD", "TRUE", "32.5580", "35.8530", "100", "FALSE",
         "Yarmouk@2025", "موظف ميداني"],
    ]

    for i, emp in enumerate(employees):
        row = start_row + i
        for j, val in enumerate(emp, 1):
            ws.cell(row=row, column=j, value=val)

    out_path = path.replace(".xlsx", "_test_filled.xlsx")
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    p = fill_test_data()
    print(f"Filled: {p}")
