# -*- coding: utf-8 -*-
"""Employee Bulk Import Template — 100% matches DB columns.

DB Model: app\models\organization.py → Employee table
API Model: app\routers\employees.py → EmployeeCreate

Column mapping (Excel col → DB field):
  A  employee_number       → employee_number        (String, UNIQUE, required)
  B  full_name             → full_name              (String, required)
  C  full_name_en          → full_name_en           (String, optional)
  D  job_title             → job_title              (String, optional)
  E  phone                 → phone                  (String, optional)
  F  email                 → email                  (String, optional)
  G  hire_date             → hire_date              (Date, optional)
  H  role                  → role_name              (String: general_manager|hr_manager|field_supervisor|office_supervisor|employee)
  I  work_type             → work_type.type_name    (String: FIELD|OFFICE|HYBRID) → auto lookup to work_type_id
  J  org_unit_name         → organization_units.unit_name (String) → auto create or find org_unit_id
  K  org_unit_type         → organization_units.unit_type  (String: COMPANY|PROVINCE|DIRECTORATE|SECTION|SUBSECTION|UNIT)
  L  direct_manager_num    → employees.employee_number → resolve to direct_manager_id (FK)
  M  work_type             → (same as I — duplicate, skip in import)
  N  allow_tracking        → allow_field_tracking   (Boolean: TRUE|FALSE)
  O  geofence_lat          → geofence_lat           (Float, optional)
  P  geofence_lng          → geofence_lng           (Float, optional)
  Q  geofence_radius       → geofence_radius_m      (Integer, default 200)
  R  geofence_exempt       → geofence_exempt        (Boolean: TRUE|FALSE)
  S  password              → password_hash          (String, hashed before save)

Hierarchy → Role mapping (auto-assigned by bulk import):
  Level 1 → general_manager
  Level 2 → general_manager
  Level 3 → hr_manager
  Level 4 → hr_manager
  Level 5 → field_supervisor
  Level 6 → office_supervisor
  Level 7 → office_supervisor
  Level 8 → employee
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

COLUMNS = [
    ("A", "employee_number",      "رقم الموظف *",           16, "فريد، مثال: EMP001"),
    ("B", "full_name",            "الاسم الكامل *",          28, "مثال: أحمد محمود عبدالله"),
    ("C", "full_name_en",         "الاسم بالإنجليزية",        22, "مثال: Ahmed Mahmoud"),
    ("D", "job_title",            "المسمى الوظيفي *",        24, "مثال: مدير مياه المحافظة"),
    ("E", "phone",                "رقم الهاتف",              14, "0791234567"),
    ("F", "email",                "البريد الإلكتروني",        26, "ahmed@yarmouk.gov.jo"),
    ("G", "hire_date",            "تاريخ التعيين",           14, "YYYY-MM-DD"),
    ("H", "role",                 "الدور الوظيفي *",          22, "general_manager|hr_manager|field_supervisor|office_supervisor|employee"),
    ("I", "org_unit_name",        "اسم الوحدة التنظيمية *",  26, "مثال: محافظة إربد / مديرية مياه الشمال"),
    ("J", "org_unit_type",        "نوع الوحدة *",           16, "COMPANY|PROVINCE|DIRECTORATE|SECTION|SUBSECTION|UNIT"),
    ("K", "direct_manager_num",   "رقم المدير المباشر *",   18, "رقم الموظف (اترك فارغاً للمدير العام فقط)"),
    ("L", "work_type",            "نوع العمل *",             14, "FIELD|OFFICE|HYBRID"),
    ("M", "allow_tracking",       "يسمح بتتبع GPS",          16, "TRUE|FALSE (مطلوب للميداني)"),
    ("N", "geofence_lat",         "خط العرض (GPS)",          14, "32.5550 (للميداني فقط)"),
    ("O", "geofence_lng",         "خط الطول (GPS)",          14, "35.8500 (للميداني فقط)"),
    ("P", "geofence_radius",      "نصف قطر المسموح (م)",    16, "200 (للميداني فقط)"),
    ("Q", "geofence_exempt",      "معفى من التقييد",        16, "TRUE|FALSE"),
    ("R", "password",             "كلمة المرور *",           20, "8+ حروف وأرقام، مثال: Yarmouk@2025"),
    ("S", "notes",                "ملاحظات",                28, "أي ملاحظات"),
]

ORG_UNIT_TYPES = ["COMPANY", "PROVINCE", "DIRECTORATE", "SECTION", "SUBSECTION", "UNIT"]
WORK_TYPES = ["FIELD", "OFFICE", "HYBRID"]
ROLES = ["general_manager", "hr_manager", "field_supervisor", "office_supervisor", "employee"]

LEVEL_TO_ROLE = {
    1: "general_manager",
    2: "general_manager",
    3: "hr_manager",
    4: "hr_manager",
    5: "field_supervisor",
    6: "office_supervisor",
    7: "office_supervisor",
    8: "employee",
}


def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _border():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _set(ws, row, col, val, bold=False, size=10, color="000000", bg=None, wrap=False, align="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", size=size, bold=bold, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = _fill(bg)
    c.border = _border()
    return c


def _title_bar(ws, row, text, cols, bg="1F3864", size=14, height=40):
    ws.row_dimensions[row].height = height
    last = get_column_letter(cols)
    ws.merge_cells(f"A{row}:{last}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = Font(name="Arial", size=size, bold=True, color="FFFFFF")
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")


# ─── Sheet 1: Instructions ───────────────────────────────────────────────────
def _sheet_instructions(wb):
    ws = wb.create_sheet("1-Instructions")
    ws.sheet_view.showGridLines = False

    widths = [4, 42, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _title_bar(ws, 1, "Employee Bulk Import — Instructions & Rules", 3)

    rows = [
        ("", None, None),
        ("STEP-BY-STEP ENTRY (Top-Down Order)", "2E75B6", True),
        ("Level 1 →  EMP001  المدير العام          | Role: general_manager | Org: COMPANY | Manager: NONE", None),
        ("Level 2 →  EMP002  نائب المدير العام     | Role: general_manager | Org: COMPANY | Manager: EMP001", None),
        ("Level 3 →  EMP003+ مديرو المحافظات       | Role: hr_manager     | Org: PROVINCE   | Manager: EMP001/EMP002", None),
        ("Level 4 →  EMPnnn+ مديرو المديريات       | Role: field_supervisor | Org: DIRECTORATE | Manager: EMP003+", None),
        ("Level 5 →  EMPnnn+ رؤساء الأقسام        | Role: office_supervisor | Org: SECTION | Manager: EMPnnn", None),
        ("Level 6 →  EMPnnn+ رؤساء الشعب          | Role: office_supervisor | Org: SUBSECTION | Manager: EMPnnn", None),
        ("Level 7 →  EMPnnn+ الموظفون             | Role: employee      | Org: UNIT       | Manager: EMPnnn", None),
        ("", None, None),
        ("REQUIRED COLUMNS", "2E75B6", True),
        ("A  employee_number    — Unique ID, e.g. EMP001", None),
        ("B  full_name          — Full name in Arabic", None),
        ("H  role              — general_manager | hr_manager | field_supervisor | office_supervisor | employee", None),
        ("D  job_title         — Job title, e.g. مدير محافظة", None),
        ("I  org_unit_name     — Organization unit name", None),
        ("J  org_unit_type     — COMPANY | PROVINCE | DIRECTORATE | SECTION | SUBSECTION | UNIT", None),
        ("K  direct_manager_num — Manager's employee_number (leave blank for Level 1 only)", None),
        ("L  work_type         — FIELD | OFFICE | HYBRID", None),
        ("M  allow_tracking    — TRUE | FALSE", None),
        ("R  password          — Min 8 chars with letters and numbers", None),
        ("", None, None),
        ("HOW THE BULK IMPORT WORKS", "2E75B6", True),
        ("1. Upload this Excel to POST /employees/bulk-import", None),
        ("2. System resolves direct_manager_num → direct_manager_id (FK lookup)", None),
        ("3. System creates or finds org_unit by name → org_unit_id (FK)", None),
        ("4. System maps work_type (FIELD/OFFICE/HYBRID) → work_type_id (FK)", None),
        ("5. System maps role name → role_id (FK)", None),
        ("6. System creates User + UserRole for each employee", None),
        ("7. Returns: created count, errors, warnings", None),
        ("", None, None),
        ("EXAMPLE (3 employees)", "2E75B6", True),
        ("EMP001 | أحمد العام | general_manager | COMPANY | — | FIELD | Yarmouk@2025", None),
        ("EMP002 | فاطمةHR | hr_manager | PROVINCE | EMP001 | OFFICE | Yarmouk@2025", None),
        ("EMP003 | خالد ميداني | field_supervisor | DIRECTORATE | EMP002 | FIELD | Yarmouk@2025", None),
    ]

    for i, row_data in enumerate(rows, 3):
        if len(row_data) == 3:
            text, bg, bold_flag = row_data
        else:
            text, bg = row_data
            bold_flag = False

        ws.row_dimensions[i].height = 22
        ws.merge_cells(f"A{i}:C{i}")
        c = ws[f"A{i}"]
        c.value = text
        if bg and bold_flag:
            c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            c.fill = _fill(bg)
        elif text.startswith("EMP") or "→" in text:
            c.font = Font(name="Courier New", size=9, color="1F3864")
            c.fill = _fill("EBF3FB")
        elif text:
            c.font = Font(name="Arial", size=10, color="404040")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    return ws


# ─── Sheet 2: Org Unit Types ────────────────────────────────────────────────
def _sheet_org_types(wb):
    ws = wb.create_sheet("2-OrgUnitTypes")
    ws.sheet_view.showGridLines = False

    widths = [5, 20, 20, 14, 20, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _title_bar(ws, 1, "Organization Unit Types Reference", 6)

    headers = ["#", "Type (EN)", "النوع (AR)", "Level", "من يحتاجه", "الوصف"]
    ws.row_dimensions[2].height = 28
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = _fill("2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()

    data = [
        (1, "COMPANY",    "الشركة",      "1-2", "المدير العام + نائب المدير", "كامل الشركة"),
        (2, "PROVINCE",   "المحافظة",    "3",   "مدير المحافظة/الإدارة",     "كل مديريات المحافظة"),
        (3, "DIRECTORATE", "المديرية",   "4",   "مدير المديرية",            "كل أقسام المديرية"),
        (4, "SECTION",    "القسم",       "5",   "رئيس القسم",               "كل شعب القسم"),
        (5, "SUBSECTION", "الشعبة",      "6",   "رئيس الشعبة",              "الموظفون في الشعبة"),
        (6, "UNIT",       "وحدة عمل",    "7-8", "الموظفون",                 "الوحدة التنظيمية للموظف"),
    ]

    bgs = ["FFF2CC", "E2EFDA", "DEEBF7", "BDD7EE", "D9E1F2", "F2F2F2"]
    for ri, (num, en, ar, lvl, who, desc) in enumerate(data, 3):
        ws.row_dimensions[ri].height = 26
        for col, val in enumerate([str(num), en, ar, lvl, who, desc], 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = _fill(bgs[ri - 3])
            c.alignment = Alignment(horizontal="center" if col < 5 else "left", vertical="center")
            c.border = _border()

    return ws


# ─── Sheet 3: Roles & Levels ────────────────────────────────────────────────
def _sheet_roles_levels(wb):
    ws = wb.create_sheet("3-RolesLevels")
    ws.sheet_view.showGridLines = False

    widths = [5, 12, 20, 20, 22, 16, 16, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _title_bar(ws, 1, "Roles & Hierarchy Levels — 7 Levels", 8)

    headers = ["#", "Level", "Role Code", "Arabic Title", "Scope", "Max Count", "Org Type", "Who Has This"]
    ws.row_dimensions[2].height = 28
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = _fill("2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()

    data = [
        (1, "1", "general_manager",    "المدير العام",           "الشركة",        "~1",    "COMPANY",    "1 person only"),
        (2, "2", "general_manager",    "نائب المدير العام",      "الشركة",        "~1-2",  "COMPANY",    "Max 2 people"),
        (3, "3", "hr_manager",         "مدير المحافظة",          "المحافظة",      "~15",   "PROVINCE",   "1 per governorate"),
        (4, "4", "field_supervisor",    "مدير المديرية",          "المديرية",     "~10+",  "DIRECTORATE","1 per directorate"),
        (5, "5", "office_supervisor",  "رئيس قسم",               "القسم",         "~3+",   "SECTION",    "1-3 per directorate"),
        (6, "6", "office_supervisor",  "رئيس شعبة",              "الشعبة",        "~3+",   "SUBSECTION", "1-3 per section"),
        (7, "7", "employee",           "موظف",                   "—",             "∞",     "UNIT",       "All remaining"),
    ]

    bgs = ["1F3864", "2E75B6", "1E6091", "3A8CC7", "5B9BD5", "7BACDD", "A3C9E5"]
    for ri, row_data in enumerate(data, 3):
        ws.row_dimensions[ri].height = 28
        bg = bgs[ri - 3]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.font = Font(name="Arial", size=10, bold=(col == 1), color="FFFFFF")
            c.fill = _fill(bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border()

    return ws


# ─── Sheet 4: Data Entry ────────────────────────────────────────────────────
def _sheet_data(wb):
    ws = wb.create_sheet("4-DataEntry")
    ws.sheet_view.showGridLines = False

    num_cols = len(COLUMNS)
    last_col = get_column_letter(num_cols)

    _title_bar(ws, 1, "Employee Data Entry — Enter TOP-DOWN: Level 1 first, then 2, then 3...", num_cols, bg="1F3864", size=12)

    ws.row_dimensions[2].height = 18
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = ("💡 Required = dark blue | Fill Level 1 first (no manager) | Level 2+ must reference manager's employee_number | "
               "Use dropdown menus | Rows 5-54 are data rows")
    c.font = Font(name="Arial", size=9, color="5C6B82")
    c.fill = _fill("EBF3FB")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[3].height = 14
    ws.merge_cells(f"A3:{last_col}3")
    c = ws["A3"]
    c.value = ("LEVEL GUIDE: 1=Director | 2=Deputy | 3=Gov.Director | 4=Dir.Director | 5=SectionHead | 6=SubSect.Chief | 7=Employee")
    c.font = Font(name="Arial", size=9, color="FFFFFF", bold=True)
    c.fill = _fill("2E75B6")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[4].height = 36
    for col_idx, (col_letter, field, header, width, hint) in enumerate(COLUMNS, 1):
        required = header.endswith(" *")
        bg = "1F3864" if required else "2E75B6"
        c = ws.cell(row=4, column=col_idx, value=header)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = _fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[5].height = 24
    for col_idx, (_, _, _, _, hint) in enumerate(COLUMNS, 1):
        c = ws.cell(row=5, column=col_idx, value=hint)
        c.font = Font(name="Arial", size=9, color="808080")
        c.fill = _fill("F2F2F2")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()

    start_row = 6
    for row in range(start_row, start_row + 50):
        ws.row_dimensions[row].height = 22
        for col_idx in range(1, num_cols + 1):
            c = ws.cell(row=row, column=col_idx)
            c.border = _border()
            c.alignment = Alignment(horizontal="center", vertical="center")
            if row % 2 == 0:
                c.fill = _fill("F7FBFF")

    ws.freeze_panes = f"A{start_row}"

    # Data Validations
    dv_role = _dv("list", '"general_manager,hr_manager,field_supervisor,office_supervisor,employee"',
                  "H", start_row, 50, "general_manager | hr_manager | field_supervisor | office_supervisor | employee",
                  "الدور الوظيفي")
    ws.add_data_validation(dv_role)

    dv_unit = _dv("list", '"COMPANY,PROVINCE,DIRECTORATE,SECTION,SUBSECTION,UNIT"',
                  "J", start_row, 50, "COMPANY | PROVINCE | DIRECTORATE | SECTION | SUBSECTION | UNIT",
                  "نوع الوحدة التنظيمية")
    ws.add_data_validation(dv_unit)

    dv_work = _dv("list", '"FIELD,OFFICE,HYBRID"',
                  "L", start_row, 50, "FIELD | OFFICE | HYBRID",
                  "نوع العمل")
    ws.add_data_validation(dv_work)

    dv_bool = _dv("list", '"TRUE,FALSE"',
                  "M", start_row, 50, "TRUE | FALSE", "GPS Tracking")
    ws.add_data_validation(dv_bool)

    dv_exempt = _dv("list", '"TRUE,FALSE"',
                    "Q", start_row, 50, "TRUE | FALSE", "Geofence Exempt")
    ws.add_data_validation(dv_exempt)

    return ws


def _dv(type_, formula, col, start, rows, prompt, title):
    return DataValidation(
        type=type_,
        formula1=formula,
        allow_blank=True,
        showErrorMessage=True,
        error="اختر من القائمة",
        errorTitle="غير صالح",
        showInputMessage=True,
        prompt=prompt,
        promptTitle=title,
    )


# ─── Sheet 5: DB Mapping ─────────────────────────────────────────────────────
def _sheet_mapping(wb):
    ws = wb.create_sheet("5-DBMapping")
    ws.sheet_view.showGridLines = False

    widths = [5, 18, 14, 28, 14, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _title_bar(ws, 1, "Column → Database Mapping (How Bulk Import Works)", 6)

    headers = ["#", "DB Column", "Type", "Description", "Required", "Import Logic"]
    ws.row_dimensions[2].height = 28
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = _fill("2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()

    mappings = [
        ("A", "employee_number", "String(20)", "Unique employee ID", "YES", "Direct insert — must be unique"),
        ("B", "full_name", "String(200)", "Full name Arabic", "YES", "Direct insert"),
        ("C", "full_name_en", "String(200)", "Full name English", "NO", "Direct insert (nullable)"),
        ("D", "job_title", "String(100)", "Job title", "YES", "Direct insert"),
        ("E", "phone", "String(20)", "Phone number", "NO", "Direct insert (nullable)"),
        ("F", "email", "String(100)", "Email address", "NO", "Direct insert (nullable)"),
        ("G", "hire_date", "Date", "Hiring date", "NO", "Direct insert (nullable)"),
        ("H", "role_name", "FK→roles.role_id", "Role name", "YES", "Lookup role by name → get role_id"),
        ("I", "org_unit_name", "FK→org_units.org_unit_id", "Org unit name", "YES", "Find or CREATE org_unit → get org_unit_id"),
        ("J", "org_unit_type", "FK→org_units.unit_type", "Unit type", "YES", "Set on org_unit (create or existing)"),
        ("K", "direct_manager_num", "FK→employees.employee_id", "Manager's emp number", "YES*", "Lookup employee_number → get employee_id → set as direct_manager_id"),
        ("L", "work_type", "FK→work_types.work_type_id", "Work type", "YES", "Lookup type_name → get work_type_id"),
        ("M", "allow_field_tracking", "Boolean", "GPS tracking allowed", "YES", "TRUE→True, FALSE→False"),
        ("N", "geofence_lat", "Float", "Latitude", "NO", "Direct insert (nullable)"),
        ("O", "geofence_lng", "Float", "Longitude", "NO", "Direct insert (nullable)"),
        ("P", "geofence_radius_m", "Integer", "Geofence radius meters", "NO", "Default 200 if blank"),
        ("Q", "geofence_exempt", "Boolean", "Exempt from geofence", "NO", "TRUE→True, FALSE→False"),
        ("R", "password_hash", "String(255)", "Password (hashed)", "YES", "hash_password() → password_hash"),
        ("S", "notes", "String(500)", "Notes", "NO", "Ignored by API — for your reference only"),
    ]

    bgs = ["F2F2F2", "FFFFFF"]
    for ri, (col, db_col, dtype, desc, req, logic) in enumerate(mappings, 3):
        ws.row_dimensions[ri].height = 26
        bg = bgs[ri % 2]
        for col_idx, val in enumerate([col, db_col, dtype, desc, req, logic], 1):
            c = ws.cell(row=ri, column=col_idx, value=val)
            c.font = Font(name="Arial", size=10, bold=(col_idx == 1))
            c.fill = _fill(bg)
            c.alignment = Alignment(horizontal="center" if col_idx <= 5 else "left", vertical="center", wrap_text=True)
            c.border = _border()

    return ws


# ─── Main ─────────────────────────────────────────────────────────────────────
def generate(output_dir: str = None) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_instructions(wb)
    _sheet_org_types(wb)
    _sheet_roles_levels(wb)
    _sheet_data(wb)
    _sheet_mapping(wb)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = f"employee_bulk_import_{ts}.xlsx"
    path = os.path.join(output_dir, name) if output_dir else name
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return path


if __name__ == "__main__":
    out = os.path.join(os.path.dirname(__file__), "output")
    p = generate(out)
    print(f"Saved: {p}")
