"""Comprehensive endpoints for the Flutter mobile app."""
from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc, delete
from typing import Optional
from datetime import datetime, date
from pydantic import BaseModel
import secrets
import io
import csv

from app.database import get_db
from app.middleware.auth import get_current_user
from app.models.organization import Employee
from app.models import (
    LeaveRequest, ShortLeave, Attendance, FieldTrackingSession, FieldTrackingPoint,
    GeofenceBreach, OvertimeWorkRequest, OvertimeWorkReport,
    ViolationNotice, Notification, Report,
    MaintenanceComplaint, Employee as EmployeeModel,
    OrganizationUnit
)

router = APIRouter(tags=["flutter_compat"])


# ==================== GPS ENDPOINTS ====================

import math as _math

GPS_TRACKER_ROLES = {"general_manager", "hr_manager"}
GPS_VIEWER_ROLES = {"general_manager", "hr_manager", "office_supervisor"}


async def _get_user_role_names(db, employee_id):
    from app.services.notifications import get_user_role_names
    return await get_user_role_names(db, employee_id)


class GpsStartRequest(BaseModel):
    target_employee_id: int
    mode: str
    interval: int
    tracking_type: str = "FIELD_WORK"
    track_color: Optional[str] = None


class GpsPointRequest(BaseModel):
    session_id: int
    latitude: float
    longitude: float
    accuracy: Optional[float] = None
    battery_level: Optional[int] = None


def _haversine_m(lat1, lon1, lat2, lon2):
    R = 6371000
    p = _math.pi / 180
    a = 0.5 - _math.cos((lat2 - lat1) * p) / 2 + _math.cos(lat1 * p) * _math.cos(lat2 * p) * (1 - _math.cos((lon2 - lon1) * p)) / 2
    return 2 * R * _math.asin(_math.sqrt(a))


@router.post("/gps/start")
async def gps_start(req: GpsStartRequest, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    roles = set(await _get_user_role_names(db, current_user.employee_id))
    if not roles.intersection(GPS_TRACKER_ROLES):
        raise HTTPException(status_code=403, detail="فقط المدير أو الموارد البشرية يمكنهم بدء التتبع")
    session = FieldTrackingSession(
        employee_id=req.target_employee_id, started_by_id=current_user.employee_id,
        tracking_type=req.tracking_type, track_mode=req.mode, track_interval=req.interval,
        started_at=datetime.utcnow(), status="active", track_color=req.track_color,
    )
    db.add(session); await db.commit(); await db.refresh(session)
    return {"session_id": session.session_id, "is_active": True, "status": session.status}


@router.post("/gps/stop")
async def gps_stop(body: dict = {}, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    roles = set(await _get_user_role_names(db, current_user.employee_id))
    session_id = body.get("session_id")
    if session_id:
        if not roles.intersection(GPS_TRACKER_ROLES):
            raise HTTPException(status_code=403, detail="فقط المدير أو الموارد البشرية يمكنهم إيقاف التتبع")
        result = await db.execute(select(FieldTrackingSession).where(FieldTrackingSession.session_id == session_id))
        session = result.scalar_one_or_none()
    else:
        result = await db.execute(
            select(FieldTrackingSession).where(
                FieldTrackingSession.employee_id == current_user.employee_id,
                FieldTrackingSession.status == "active"
            )
        )
        session = result.scalar_one_or_none()
    if not session:
        return {"status": "no_active_session"}
    session.status = "ended"; session.ended_at = datetime.utcnow()
    await db.commit()
    return {"session_id": session.session_id, "status": "stopped"}


@router.post("/gps/point")
async def gps_point(req: GpsPointRequest, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(
        select(FieldTrackingSession).where(FieldTrackingSession.session_id == req.session_id)
    )
    session = result.scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    point = FieldTrackingPoint(
        session_id=req.session_id, employee_id=current_user.employee_id,
        latitude=req.latitude, longitude=req.longitude, accuracy=req.accuracy,
        battery_level=req.battery_level, recorded_at=datetime.utcnow(),
    )
    db.add(point)

    session.last_lat = req.latitude
    session.last_lng = req.longitude
    session.last_point_at = datetime.utcnow()

    emp_res = await db.execute(select(Employee).where(Employee.employee_id == session.employee_id))
    emp = emp_res.scalar_one_or_none()
    if emp and emp.geofence_lat and emp.geofence_lng:
        dist = _haversine_m(req.latitude, req.longitude, emp.geofence_lat, emp.geofence_lng)
        radius = emp.geofence_radius_m or 200
        was_outside = session.is_outside
        if dist > radius:
            session.is_outside = True
            session.outside_distance_m = dist - radius
            if not was_outside:
                session.outside_started_at = datetime.utcnow()
        else:
            if was_outside and session.outside_started_at:
                breach = GeofenceBreach(
                    session_id=session.session_id, employee_id=session.employee_id,
                    started_at=session.outside_started_at, ended_at=datetime.utcnow(),
                    duration_seconds=(datetime.utcnow() - session.outside_started_at).total_seconds(),
                    distance_m=session.outside_distance_m,
                    start_latitude=session.last_lat, start_longitude=session.last_lng,
                    end_latitude=req.latitude, end_longitude=req.longitude,
                )
                db.add(breach)
            session.is_outside = False
            session.outside_distance_m = 0.0
            session.outside_started_at = None

    await db.commit(); await db.refresh(point)
    return {"point_id": point.point_id, "status": "recorded", "is_outside": session.is_outside}


@router.get("/gps/viewer")
async def gps_viewer(db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(
        select(FieldTrackingSession).where(
            FieldTrackingSession.viewer_employee_id == current_user.employee_id,
            FieldTrackingSession.status == "active"
        ).limit(1)
    )
    session = result.scalar_one_or_none()
    if session:
        emp_res = await db.execute(select(Employee).where(Employee.employee_id == session.started_by_id))
        emp = emp_res.scalar_one_or_none()
        return {
            "is_viewer": True,
            "viewer": {"full_name": emp.full_name if emp else "", "employee_id": session.started_by_id}
        }
    return {"is_viewer": False, "viewer": None}


@router.post("/gps/set-viewer")
async def gps_set_viewer(body: dict, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    roles = set(await _get_user_role_names(db, current_user.employee_id))
    if not roles.intersection(GPS_TRACKER_ROLES):
        raise HTTPException(status_code=403, detail="فقط المدير أو الموارد البشرية يمكنهم تعيين متابع")
    viewer_id = body.get("employee_id")
    result = await db.execute(
        select(FieldTrackingSession).where(
            FieldTrackingSession.started_by_id == current_user.employee_id,
            FieldTrackingSession.status == "active"
        )
    )
    sessions = result.scalars().all()
    for s in sessions:
        s.viewer_employee_id = viewer_id
    await db.commit()
    return {"status": "updated", "viewer_employee_id": viewer_id}


@router.get("/gps/history")
async def gps_history(employee_id: int, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(
        select(FieldTrackingSession).where(FieldTrackingSession.employee_id == employee_id)
        .order_by(desc(FieldTrackingSession.started_at)).limit(50)
    )
    sessions = result.scalars().all()
    return {"sessions": [
        {"session_id": s.session_id, "started_at": s.started_at.isoformat() if s.started_at else None,
         "ended_at": s.ended_at.isoformat() if s.ended_at else None, "status": s.status,
         "track_color": s.track_color}
        for s in sessions
    ]}


@router.post("/gps/simulate-point")
async def gps_simulate_point(body: dict, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    roles = set(await _get_user_role_names(db, current_user.employee_id))
    if not roles.intersection(GPS_TRACKER_ROLES):
        raise HTTPException(status_code=403, detail="فقط المدير أو الموارد البشرية يمكنهم المحاكاة")
    point = FieldTrackingPoint(
        session_id=body.get("session_id"), employee_id=current_user.employee_id,
        latitude=body.get("latitude", 0), longitude=body.get("longitude", 0),
        recorded_at=datetime.utcnow(),
    )
    db.add(point); await db.commit()
    return {"status": "simulated", "point_id": point.point_id}


# ==================== LEAVE ====================

class LeaveReviewRequest(BaseModel):
    status: str
    review_note: Optional[str] = None


class ShortLeaveRequest(BaseModel):
    leave_kind: str
    outing_date: str
    departure_time: str
    return_time: str
    destination: Optional[str] = None
    reason: Optional[str] = None
    tracking_required: bool = False
    tracking_acknowledged: bool = False


@router.post("/leave/{request_id}/review")
async def review_leave(request_id: int, req: LeaveReviewRequest, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(select(LeaveRequest).where(LeaveRequest.request_id == request_id))
    leave = result.scalar_one_or_none()
    if not leave:
        raise HTTPException(status_code=404, detail="Leave request not found")
    leave.status = req.status
    leave.approved_by = current_user.employee_id
    leave.review_note = req.review_note
    leave.reviewed_at = datetime.utcnow()
    await db.commit()

    from app.services.notifications import notify_employee
    status_label = "تم اعتماد" if req.status == "approved" else "تم رفض"
    await notify_employee(
        db,
        leave.employee_id,
        f"{status_label} طلب الإجازة",
        f" {status_label} طلب الإجازة رقم {request_id} من قبل {current_user.full_name}",
        severity="success" if req.status == "approved" else "danger",
    )

    return {"status": leave.status, "request_id": request_id}


@router.post("/leave/short")
async def create_short_leave(req: ShortLeaveRequest, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    from datetime import time as dt_time
    dep_h, dep_m = map(int, req.departure_time.split(":"))
    ret_h, ret_m = map(int, req.return_time.split(":"))
    short_leave = ShortLeave(
        employee_id=current_user.employee_id,
        leave_kind=req.leave_kind,
        outing_date=datetime.fromisoformat(req.outing_date).date() if req.outing_date else date.today(),
        departure_time=dt_time(dep_h, dep_m),
        return_time=dt_time(ret_h, ret_m),
        destination=req.destination,
        reason=req.reason,
        tracking_required=req.tracking_required,
        tracking_acknowledged=req.tracking_acknowledged,
        status="pending",
    )
    db.add(short_leave)
    await db.commit()
    await db.refresh(short_leave)

    from app.services.notifications import notify_direct_manager
    kind_label = "رسمية" if req.leave_kind == "official" else "خاصة"
    await notify_direct_manager(
        db,
        current_user.employee_id,
        "طلب مغادرة جديد",
        f" الموظف {current_user.full_name} تقدم بطلب مغادرة {kind_label} في {req.outing_date} من {req.departure_time} إلى {req.return_time}",
        severity="info",
    )

    return {"short_leave_id": short_leave.short_leave_id, "status": short_leave.status, "leave_kind": req.leave_kind}


@router.get("/leave/short/my")
async def my_short_leaves(db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(
        select(ShortLeave)
        .where(ShortLeave.employee_id == current_user.employee_id)
        .order_by(desc(ShortLeave.created_at))
        .limit(100)
    )
    leaves = result.scalars().all()
    return {
        "items": [
            {
                "short_leave_id": s.short_leave_id,
                "leave_kind": s.leave_kind,
                "outing_date": s.outing_date.isoformat() if s.outing_date else None,
                "departure_time": s.departure_time,
                "return_time": s.return_time,
                "destination": s.destination,
                "reason": s.reason,
                "tracking_required": s.tracking_required,
                "tracking_acknowledged": s.tracking_acknowledged,
                "status": s.status,
                "employee_name": current_user.full_name,
            }
            for s in leaves
        ]
    }


@router.post("/leave/short/{short_leave_id}/review")
async def review_short_leave(short_leave_id: int, req: LeaveReviewRequest, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(select(ShortLeave).where(ShortLeave.short_leave_id == short_leave_id))
    leave = result.scalar_one_or_none()
    if not leave:
        raise HTTPException(status_code=404, detail="Short leave not found")
    leave.status = req.status
    leave.approved_by = current_user.employee_id
    leave.review_note = req.review_note
    leave.reviewed_at = datetime.utcnow()
    await db.commit()

    from app.services.notifications import notify_employee
    status_label = "تم اعتماد" if req.status == "approved" else "تم رفض"
    await notify_employee(
        db,
        leave.employee_id,
        f"{status_label} طلب المغادرة",
        f" {status_label} طلب المغادرة رقم {short_leave_id} من قبل {current_user.full_name}",
        severity="success" if req.status == "approved" else "danger",
    )

    return {"status": leave.status, "short_leave_id": short_leave_id}


# ==================== OVERTIME ====================

class OvertimeCreateRequest(BaseModel):
    task_description: str
    area_name: str
    area_lat: float
    area_lng: float
    requested_hours: float
    work_date: Optional[str] = None
    work_type: Optional[str] = "field"


class OvertimeReviewRequest(BaseModel):
    status: str
    total_approved_hours: Optional[float] = None
    review_note: Optional[str] = None


class OvertimeExtendRequest(BaseModel):
    additional_hours: float
    review_note: Optional[str] = None


class OvertimeReportRequest(BaseModel):
    work_done: str
    actual_hours: Optional[float] = None
    actual_lat: Optional[float] = None
    actual_lng: Optional[float] = None
    photo_url: Optional[str] = None


@router.post("/overtime-work/")
async def create_overtime(req: OvertimeCreateRequest, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    work_date_val = datetime.strptime(req.work_date, "%Y-%m-%d").date() if req.work_date else date.today()
    ot = OvertimeWorkRequest(
        employee_id=current_user.employee_id,
        work_date=work_date_val,
        work_type=req.work_type or "field",
        task_description=req.task_description,
        area_name=req.area_name, area_lat=req.area_lat, area_lng=req.area_lng,
        requested_hours=req.requested_hours, status="pending",
    )
    db.add(ot); await db.commit(); await db.refresh(ot)

    from app.services.notifications import notify_direct_manager
    work_type_labels = {"field": "ميداني", "office": "مكتبي", "maintenance": "صيانة", "other": "أخرى"}
    work_type_label = work_type_labels.get(req.work_type or "field", req.work_type or "ميداني")
    await notify_direct_manager(
        db,
        current_user.employee_id,
        "طلب عمل إضافي جديد",
        f" الموظف {current_user.full_name} تقدم بطلب عمل إضافي ({work_type_label}) في {req.area_name} لمدة {req.requested_hours} ساعات",
        severity="info",
    )

    return {"request_id": ot.request_id, "status": ot.status}


@router.get("/overtime-work/my")
async def my_overtime(db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(
        select(OvertimeWorkRequest).where(OvertimeWorkRequest.employee_id == current_user.employee_id)
        .order_by(desc(OvertimeWorkRequest.created_at)).limit(50)
    )
    requests = result.scalars().all()
    return {"items": [
        {
            "request_id": r.request_id,
            "employee_id": r.employee_id,
            "employee_number": current_user.employee_number,
            "employee_name": current_user.full_name,
            "work_date": r.work_date.isoformat() if r.work_date else None,
            "work_type": r.work_type,
            "task_description": r.task_description,
            "area_name": r.area_name,
            "area_lat": r.area_lat,
            "area_lng": r.area_lng,
            "requested_hours": r.requested_hours,
            "extended_hours": r.extended_hours,
            "total_approved_hours": r.total_approved_hours,
            "actual_hours": r.actual_hours,
            "status": r.status,
            "tracking_session_id": r.tracking_session_id,
            "tracking_starts_at": r.tracking_starts_at.isoformat() if r.tracking_starts_at else None,
            "tracking_ends_at": r.tracking_ends_at.isoformat() if r.tracking_ends_at else None,
            "completed_at": r.completed_at.isoformat() if r.completed_at else None,
            "completed_lat": r.completed_lat,
            "completed_lng": r.completed_lng,
            "completed_photo_url": r.completed_photo_url,
            "review_note": r.review_note,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in requests
    ]}


@router.get("/overtime-work/all")
async def all_overtime(db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(select(OvertimeWorkRequest).order_by(desc(OvertimeWorkRequest.created_at)).limit(100))
    requests = result.scalars().all()
    items = []
    for r in requests:
        emp_res = await db.execute(select(Employee).where(Employee.employee_id == r.employee_id))
        emp = emp_res.scalar_one_or_none()
        items.append({
            "request_id": r.request_id,
            "employee_id": r.employee_id,
            "employee_number": emp.employee_number if emp else None,
            "employee_name": emp.full_name if emp else None,
            "work_date": r.work_date.isoformat() if r.work_date else None,
            "work_type": r.work_type,
            "task_description": r.task_description,
            "area_name": r.area_name,
            "area_lat": r.area_lat,
            "area_lng": r.area_lng,
            "requested_hours": r.requested_hours,
            "extended_hours": r.extended_hours,
            "total_approved_hours": r.total_approved_hours,
            "actual_hours": r.actual_hours,
            "status": r.status,
            "tracking_session_id": r.tracking_session_id,
            "tracking_starts_at": r.tracking_starts_at.isoformat() if r.tracking_starts_at else None,
            "tracking_ends_at": r.tracking_ends_at.isoformat() if r.tracking_ends_at else None,
            "completed_at": r.completed_at.isoformat() if r.completed_at else None,
            "completed_lat": r.completed_lat,
            "completed_lng": r.completed_lng,
            "completed_photo_url": r.completed_photo_url,
            "review_note": r.review_note,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })
    return {"items": items}


@router.post("/overtime-work/{request_id}/review")
async def review_overtime(request_id: int, req: OvertimeReviewRequest, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(select(OvertimeWorkRequest).where(OvertimeWorkRequest.request_id == request_id))
    ot = result.scalar_one_or_none()
    if not ot:
        raise HTTPException(status_code=404, detail="Overtime request not found")
    ot.status = req.status
    if req.total_approved_hours is not None:
        ot.total_approved_hours = req.total_approved_hours
    else:
        ot.total_approved_hours = ot.requested_hours + ot.extended_hours
    if req.review_note is not None:
        ot.review_note = req.review_note
    ot.reviewed_by = current_user.employee_id
    ot.reviewed_at = datetime.utcnow()
    await db.commit()

    from app.services.notifications import notify_employee
    if req.status == "approved":
        await notify_employee(
            db, ot.employee_id,
            "تم اعتماد طلب العمل الإضافي",
            f" تم اعتماد طلب العمل الإضافي رقم {request_id} لمدة {ot.total_approved_hours} ساعات. يمكنكم بدء العمل.",
            severity="success",
        )
    elif req.status == "rejected":
        await notify_employee(
            db, ot.employee_id,
            "تم رفض طلب العمل الإضافي",
            f" تم رفض طلب العمل الإضافي رقم {request_id}. السبب: {req.review_note or 'غير محدد'}",
            severity="danger",
        )

    return {"status": ot.status, "request_id": request_id}


@router.post("/overtime-work/{request_id}/extend")
async def extend_overtime(request_id: int, req: OvertimeExtendRequest, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(select(OvertimeWorkRequest).where(OvertimeWorkRequest.request_id == request_id))
    ot = result.scalar_one_or_none()
    if not ot:
        raise HTTPException(status_code=404, detail="Overtime request not found")
    ot.extended_hours = (ot.extended_hours or 0) + req.additional_hours
    ot.extended_by = current_user.employee_id
    ot.extended_at = datetime.utcnow()
    if req.review_note:
        ot.review_note = (ot.review_note or "") + f"\nتمديد: {req.review_note}"
    await db.commit()

    from app.services.notifications import notify_employee
    await notify_employee(
        db, ot.employee_id,
        "تمديد العمل الإضافي",
        f" تمت إضافة {req.additional_hours} ساعات إضافية لطلب العمل الإضافي رقم {request_id}. الساعات الإجمالية: {ot.requested_hours + ot.extended_hours}",
        severity="info",
    )

    return {"status": "extended", "new_total_hours": ot.requested_hours + ot.extended_hours}


@router.post("/overtime-work/{request_id}/complete")
async def complete_overtime(request_id: int, req: OvertimeReportRequest, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(select(OvertimeWorkRequest).where(OvertimeWorkRequest.request_id == request_id))
    ot = result.scalar_one_or_none()
    if not ot:
        raise HTTPException(status_code=404, detail="Overtime request not found")

    ot.actual_hours = req.actual_hours
    ot.completed_at = datetime.utcnow()
    ot.completed_lat = req.actual_lat
    ot.completed_lng = req.actual_lng
    ot.completed_photo_url = req.photo_url
    ot.status = "completed"

    if ot.tracking_starts_at:
        ot.tracking_ends_at = datetime.utcnow()

    report = OvertimeWorkReport(
        request_id=request_id,
        work_done=req.work_done,
        actual_hours=req.actual_hours,
        actual_lat=req.actual_lat,
        actual_lng=req.actual_lng,
        photo_url=req.photo_url,
        submitted_at=datetime.utcnow(),
    )
    db.add(report)
    await db.commit()
    await db.refresh(report)

    from app.services.notifications import notify_direct_manager
    emp_res = await db.execute(select(Employee).where(Employee.employee_id == current_user.employee_id))
    emp = emp_res.scalar_one_or_none()
    emp_name = emp.full_name if emp else current_user.full_name
    emp_num = emp.employee_number if emp else current_user.employee_number

    await notify_direct_manager(
        db,
        current_user.employee_id,
        "تقرير إنهاء العمل الإضافي",
        f" الموظف {emp_name} ({emp_num}) أنهى العمل الإضافي رقم {request_id}. الوقت الفعلي: {req.actual_hours} ساعات",
        severity="success",
    )

    return {
        "report_id": report.report_id,
        "status": "completed",
        "actual_hours": req.actual_hours,
        "completed_at": ot.completed_at.isoformat(),
    }


@router.post("/overtime-work/{request_id}/report")
async def submit_overtime_report(request_id: int, req: OvertimeReportRequest, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    report = OvertimeWorkReport(
        request_id=request_id, work_done=req.work_done,
        actual_hours=req.actual_hours, actual_lat=req.actual_lat, actual_lng=req.actual_lng,
        photo_url=req.photo_url, submitted_at=datetime.utcnow(),
    )
    db.add(report); await db.commit(); await db.refresh(report)
    return {"report_id": report.report_id, "status": "submitted"}


@router.get("/overtime-work/{request_id}/reports")
async def get_overtime_reports(request_id: int, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(select(OvertimeWorkReport).where(OvertimeWorkReport.request_id == request_id))
    reports = result.scalars().all()
    return {"items": [
        {
            "report_id": r.report_id,
            "work_done": r.work_done,
            "actual_hours": r.actual_hours,
            "actual_lat": r.actual_lat,
            "actual_lng": r.actual_lng,
            "photo_url": r.photo_url,
            "submitted_at": r.submitted_at.isoformat() if r.submitted_at else None,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in reports
    ]}


@router.get("/overtime-work/{request_id}/print")
async def print_overtime_report(request_id: int, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(select(OvertimeWorkRequest).where(OvertimeWorkRequest.request_id == request_id))
    ot = result.scalar_one_or_none()
    if not ot:
        raise HTTPException(status_code=404, detail="Overtime request not found")

    emp_res = await db.execute(select(Employee).where(Employee.employee_id == ot.employee_id))
    emp = emp_res.scalar_one_or_none()

    report_result = await db.execute(select(OvertimeWorkReport).where(OvertimeWorkReport.request_id == request_id))
    reports = report_result.scalars().all()

    reviewer = None
    if ot.reviewed_by:
        rev_res = await db.execute(select(Employee).where(Employee.employee_id == ot.reviewed_by))
        reviewer = rev_res.scalar_one_or_none()

    work_type_labels = {"field": "ميداني", "office": "مكتبي", "maintenance": "صيانة", "other": "أخرى"}

    return {
        "request": {
            "request_id": ot.request_id,
            "employee_name": emp.full_name if emp else "غير معروف",
            "employee_number": emp.employee_number if emp else "غير معروف",
            "job_title": emp.job_title if emp else "غير محدد",
            "work_date": ot.work_date.isoformat() if ot.work_date else None,
            "work_type": work_type_labels.get(ot.work_type, ot.work_type),
            "task_description": ot.task_description,
            "area_name": ot.area_name,
            "requested_hours": ot.requested_hours,
            "extended_hours": ot.extended_hours,
            "total_approved_hours": ot.total_approved_hours,
            "actual_hours": ot.actual_hours,
            "status": ot.status,
            "tracking_starts_at": ot.tracking_starts_at.isoformat() if ot.tracking_starts_at else None,
            "tracking_ends_at": ot.tracking_ends_at.isoformat() if ot.tracking_ends_at else None,
            "completed_at": ot.completed_at.isoformat() if ot.completed_at else None,
            "completed_lat": ot.completed_lat,
            "completed_lng": ot.completed_lng,
            "completed_photo_url": ot.completed_photo_url,
            "reviewer_name": reviewer.full_name if reviewer else None,
            "review_note": ot.review_note,
            "created_at": ot.created_at.isoformat() if ot.created_at else None,
        },
        "reports": [
            {
                "report_id": r.report_id,
                "work_done": r.work_done,
                "actual_hours": r.actual_hours,
                "actual_lat": r.actual_lat,
                "actual_lng": r.actual_lng,
                "photo_url": r.photo_url,
                "submitted_at": r.submitted_at.isoformat() if r.submitted_at else None,
            }
            for r in reports
        ],
    }


# ==================== VIOLATIONS ====================

class ViolationCreateRequest(BaseModel):
    employee_id: int
    violation_type: str
    violation_date: str
    violation_time: str
    penalty: str
    notes: Optional[str] = None


@router.post("/violations/")
async def create_violation(req: ViolationCreateRequest, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    v = ViolationNotice(
        issuer_id=current_user.employee_id, employee_id=req.employee_id,
        violation_type=req.violation_type,
        violation_date=datetime.fromisoformat(req.violation_date).date() if req.violation_date else date.today(),
        violation_time=datetime.strptime(req.violation_time, "%H:%M").time() if ":" in str(req.violation_time) else None,
        penalty=req.penalty, notes=req.notes,
    )
    db.add(v); await db.commit(); await db.refresh(v)
    return {"violation_id": v.violation_id, "status": "created"}


@router.get("/violations/my")
async def my_violations(db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(
        select(ViolationNotice).where(ViolationNotice.employee_id == current_user.employee_id)
        .order_by(desc(ViolationNotice.created_at)).limit(50)
    )
    violations = result.scalars().all()
    return {"items": [
        {"violation_id": v.violation_id, "violation_type": v.violation_type,
         "violation_date": v.violation_date.isoformat() if v.violation_date else None,
         "penalty": v.penalty, "notes": v.notes}
        for v in violations
    ]}


@router.post("/violations/{violation_id}/acknowledge")
async def acknowledge_violation(violation_id: int, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(select(ViolationNotice).where(ViolationNotice.violation_id == violation_id))
    v = result.scalar_one_or_none()
    if not v:
        raise HTTPException(status_code=404, detail="Violation not found")
    return {"status": "acknowledged", "violation_id": violation_id, "violation": {"violation_id": v.violation_id}}


@router.post("/violations/{violation_id}/respond")
async def respond_violation(violation_id: int, body: dict, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    return {"status": "responded", "violation_id": violation_id, "response": body.get("response")}


# ==================== NOTIFICATIONS ====================

@router.post("/notifications/read-all")
async def mark_all_notifications_read(db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(select(Notification).where(Notification.employee_id == current_user.employee_id))
    notifications = result.scalars().all()
    for n in notifications:
        n.is_read = True
    await db.commit()
    return {"status": "all_read", "count": len(notifications)}


# ==================== REPORTS ====================

class ReportCreateRequest(BaseModel):
    report_type: str = "daily"
    title: str
    description: str


@router.post("/reports/")
async def create_report(req: ReportCreateRequest, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    report = Report(
        employee_id=current_user.employee_id, report_type=req.report_type,
        title=req.title, description=req.description,
        report_date=date.today(), status="pending",
    )
    db.add(report); await db.commit(); await db.refresh(report)

    from app.services.notifications import notify_direct_manager
    await notify_direct_manager(
        db,
        current_user.employee_id,
        "تقرير جديد",
        f" الموظف {current_user.full_name} أرسل تقريراً: {req.title}",
        severity="info",
    )

    return {"report_id": report.report_id, "status": report.status}


# ==================== ADMIN - EMPLOYEES ====================

class EmployeeCreateRequest(BaseModel):
    employee_number: str
    full_name: str
    full_name_en: Optional[str] = None
    job_title: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    org_unit_id: Optional[int] = None
    work_type_id: Optional[int] = None
    direct_manager_id: Optional[int] = None
    hire_date: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None  # Role name like "employee", "hr_manager", etc.


@router.post("/employees/")
async def create_employee(req: EmployeeCreateRequest, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    emp = EmployeeModel(
        employee_number=req.employee_number, full_name=req.full_name,
        full_name_en=req.full_name_en, job_title=req.job_title,
        phone=req.phone, email=req.email,
        org_unit_id=req.org_unit_id, work_type_id=req.work_type_id,
        direct_manager_id=req.direct_manager_id,
        hire_date=datetime.fromisoformat(req.hire_date).date() if req.hire_date else None,
        is_active=True,
    )
    db.add(emp); await db.commit(); await db.refresh(emp)

    if req.password:
        from app.models.auth import User
        import bcrypt
        user = User(
            username=req.employee_number,
            employee_id=emp.employee_id,
            password_hash=bcrypt.hashpw(req.password.encode(), bcrypt.gensalt()).decode(),
            is_active=True,
        )
        db.add(user); await db.commit(); await db.refresh(user)
        
        # Assign role if provided
        if req.role:
            from app.models.auth import Role, UserRole
            role_result = await db.execute(select(Role).where(Role.role_name == req.role))
            role = role_result.scalar_one_or_none()
            if role:
                user_role = UserRole(user_id=user.user_id, role_id=role.role_id)
                db.add(user_role); await db.commit()

    return {"employee_id": emp.employee_id, "status": "created"}


@router.patch("/employees/{employee_id}")
async def update_employee(employee_id: int, body: dict, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(select(EmployeeModel).where(EmployeeModel.employee_id == employee_id))
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Handle role update separately
    role_name = body.pop("role", None)
    
    for key, value in body.items():
        if hasattr(emp, key) and key not in ("employee_id", "created_at", "updated_at", "password_hash", "pattern_hash"):
            setattr(emp, key, value)
    await db.commit()
    
    # Update role if provided
    if role_name:
        from app.models.auth import User, Role, UserRole
        user_result = await db.execute(select(User).where(User.employee_id == employee_id))
        user = user_result.scalar_one_or_none()
        if user:
            # Get role
            role_result = await db.execute(select(Role).where(Role.role_name == role_name))
            role = role_result.scalar_one_or_none()
            if role:
                # Remove existing roles
                await db.execute(delete(UserRole).where(UserRole.user_id == user.user_id))
                # Add new role
                user_role = UserRole(user_id=user.user_id, role_id=role.role_id)
                db.add(user_role); await db.commit()
    
    return {"status": "updated", "employee_id": employee_id}


@router.delete("/employees/{employee_id}")
async def delete_employee(employee_id: int, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(select(EmployeeModel).where(EmployeeModel.employee_id == employee_id))
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    emp.is_active = False
    await db.commit()
    return {"status": "deactivated", "employee_id": employee_id}


@router.delete("/employees/{employee_id}/hard")
async def hard_delete_employee(employee_id: int, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(select(EmployeeModel).where(EmployeeModel.employee_id == employee_id))
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    await db.delete(emp); await db.commit()
    return {"status": "deleted", "employee_id": employee_id}


@router.post("/employees/bulk-import")
async def bulk_import_employees(file: UploadFile = File(...), db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    """Bulk import employees from Excel/CSV file."""
    content = await file.read()
    filename = file.filename or ""
    
    imported = 0
    errors = []
    
    # Arabic to English field mapping
    field_map = {
        "رقم الموظف": "employee_number",
        "الرقم الوظيفي": "employee_number",
        "الاسم الكامل": "full_name",
        "المسمى الوظيفي": "job_title",
        "الهاتف": "phone",
        "رقم الهاتف": "phone",
        "البريد الإلكتروني": "email",
        "البريد": "email",
        "الدور": "role",
        "المديرية": "directorate",
        "القسم": "department",
        "الشعبة": "section",
        "الوحدة التنظيمية": "org_unit_name",
        "الوحدة التنظيمية (كود)": "org_unit_code",
        "نوع العمل": "work_type",
        "المدير المباشر": "manager_number",
        "رقم المدير المباشر": "manager_number",
        "المدير المباشر (رقم وظيفي)": "manager_number",
        "تاريخ التعيين": "hire_date",
        "كلمة المرور": "password",
        "الاسم بالإنجليزية": "full_name_en",
        # Also accept English names
        "employee_number": "employee_number",
        "full_name": "full_name",
        "full_name_en": "full_name_en",
        "job_title": "job_title",
        "phone": "phone",
        "email": "email",
        "role": "role",
        "org_unit_id": "org_unit_id",
        "work_type_id": "work_type_id",
        "direct_manager_id": "direct_manager_id",
        "hire_date": "hire_date",
        "password": "password",
    }
    
    try:
        if filename.endswith(('.xlsx', '.xls')):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
            except ImportError:
                return {"status": "error", "message": "openpyxl not installed"}
        else:
            text = content.decode('utf-8-sig')
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
        
        if not rows:
            return {"status": "empty", "imported": 0}
        
        # Map headers to English field names
        raw_headers = [str(h).strip() if h else "" for h in rows[0]]
        headers = [field_map.get(h, field_map.get(h.lower(), h.lower())) for h in raw_headers]
        header_map = {h: i for i, h in enumerate(headers)}
        
        for i, row in enumerate(rows[1:], start=2):
            try:
                def get_val(key, default=None):
                    idx = header_map.get(key)
                    if idx is not None and idx < len(row) and row[idx] is not None:
                        return str(row[idx]).strip()
                    return default
                
                emp_num = get_val("employee_number")
                full_name = get_val("full_name")
                password = get_val("password", "Yarmouk@2025")
                role_name = get_val("role")
                org_unit_name = get_val("org_unit_name")
                work_type_name = get_val("work_type")
                manager_number = get_val("manager_number")
                
                # Get hierarchy levels
                directorate_name = get_val("directorate")
                department_name = get_val("department")
                section_name = get_val("section")
                org_unit_code = get_val("org_unit_code")
                
                if not emp_num or not full_name:
                    continue
                
                # Check if exists
                existing = await db.execute(
                    select(EmployeeModel).where(EmployeeModel.employee_number == emp_num)
                )
                if existing.scalar_one_or_none():
                    continue
                
                # Resolve org_unit_id from hierarchy (section > department > directorate > code > name)
                from app.models.organization import OrganizationUnit
                org_unit_id = None
                
                # Try org_unit_code first (most reliable)
                if org_unit_code:
                    code_result = await db.execute(
                        select(OrganizationUnit).where(OrganizationUnit.unit_code == org_unit_code)
                    )
                    code_unit = code_result.scalar_one_or_none()
                    if code_unit:
                        org_unit_id = code_unit.org_unit_id
                
                # Try section (lowest level)
                if org_unit_id is None and section_name:
                    section_result = await db.execute(
                        select(OrganizationUnit).where(
                            OrganizationUnit.unit_name == section_name,
                            OrganizationUnit.unit_type == "SECTION"
                        )
                    )
                    section = section_result.scalar_one_or_none()
                    if section:
                        org_unit_id = section.org_unit_id
                
                # Try department
                if org_unit_id is None and department_name:
                    dept_result = await db.execute(
                        select(OrganizationUnit).where(
                            OrganizationUnit.unit_name == department_name,
                            OrganizationUnit.unit_type == "DEPARTMENT"
                        )
                    )
                    dept = dept_result.scalar_one_or_none()
                    if dept:
                        org_unit_id = dept.org_unit_id
                
                # Try directorate
                if org_unit_id is None and directorate_name:
                    dir_result = await db.execute(
                        select(OrganizationUnit).where(
                            OrganizationUnit.unit_name == directorate_name,
                            OrganizationUnit.unit_type == "DIRECTORATE"
                        )
                    )
                    directorate = dir_result.scalar_one_or_none()
                    if directorate:
                        org_unit_id = directorate.org_unit_id
                
                # Fallback to org_unit_name or org_unit_id
                if org_unit_id is None and org_unit_name:
                    unit_result = await db.execute(
                        select(OrganizationUnit).where(OrganizationUnit.unit_name == org_unit_name)
                    )
                    unit = unit_result.scalar_one_or_none()
                    if unit:
                        org_unit_id = unit.org_unit_id
                    elif org_unit_name.isdigit():
                        org_unit_id = int(org_unit_name)
                
                # Resolve work_type_id from name (use database lookup, not hardcoded)
                work_type_id = None
                if work_type_name:
                    from app.models.organization import WorkType
                    wt_result = await db.execute(
                        select(WorkType).where(WorkType.type_name_ar == work_type_name)
                    )
                    wt = wt_result.scalar_one_or_none()
                    if not wt:
                        wt_result2 = await db.execute(
                            select(WorkType).where(WorkType.type_name == work_type_name)
                        )
                        wt = wt_result2.scalar_one_or_none()
                    if wt:
                        work_type_id = wt.work_type_id
                    elif work_type_name.isdigit():
                        work_type_id = int(work_type_name)
                
                # Resolve manager_id from employee_number
                direct_manager_id = None
                if manager_number:
                    manager_result = await db.execute(
                        select(EmployeeModel).where(EmployeeModel.employee_number == manager_number)
                    )
                    manager = manager_result.scalar_one_or_none()
                    if manager:
                        direct_manager_id = manager.employee_id
                    elif manager_number.isdigit():
                        direct_manager_id = int(manager_number)
                
                # Parse hire_date
                hire_date = get_val("hire_date")
                parsed_hire_date = None
                if hire_date:
                    try:
                        parsed_hire_date = datetime.strptime(hire_date, "%Y-%m-%d").date()
                    except ValueError:
                        try:
                            parsed_hire_date = datetime.strptime(hire_date, "%d/%m/%Y").date()
                        except ValueError:
                            pass
                
                emp = EmployeeModel(
                    employee_number=emp_num,
                    full_name=full_name,
                    full_name_en=get_val("full_name_en"),
                    job_title=get_val("job_title"),
                    phone=get_val("phone"),
                    email=get_val("email"),
                    org_unit_id=org_unit_id,
                    work_type_id=work_type_id,
                    direct_manager_id=direct_manager_id,
                    hire_date=parsed_hire_date,
                    is_active=True,
                )
                db.add(emp)
                await db.commit()
                await db.refresh(emp)
                
                # Create user with password
                from app.models.auth import User, Role, UserRole
                import bcrypt
                user = User(
                    username=emp_num,
                    employee_id=emp.employee_id,
                    password_hash=bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode(),
                    is_active=True,
                )
                db.add(user)
                await db.commit()
                await db.refresh(user)
                
                # Assign role if provided
                if role_name:
                    role_result = await db.execute(select(Role).where(Role.role_name == role_name))
                    role = role_result.scalar_one_or_none()
                    if role:
                        user_role = UserRole(user_id=user.user_id, role_id=role.role_id)
                        db.add(user_role)
                        await db.commit()
                
                imported += 1
            except Exception as e:
                await db.rollback()
                errors.append({"row": i, "error": str(e)[:200]})
        
        return {
            "status": "imported",
            "imported": imported,
            "total": len(rows) - 1,
            "errors": errors[:10],
            "message": f"تم استيراد {imported} من {len(rows) - 1} موظف"
        }
    except Exception as e:
        return {"status": "error", "message": str(e)[:300]}


@router.get("/employees/template")
async def download_template():
    """Download Excel template for employee bulk import."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        wb = Workbook()
        ws = wb.active
        ws.title = "الموظفين"
        
        # Arabic headers with hierarchy columns
        headers = [
            "المديرية", "القسم", "الشعبة",
            "رقم الموظف", "الاسم الكامل", "المسمى الوظيفي",
            "الهاتف", "البريد الإلكتروني", "الدور",
            "رقم المدير المباشر", "تاريخ التعيين", "كلمة المرور"
        ]
        ws.append(headers)
        
        # Sample rows showing hierarchy
        sample_rows = [
            ["مديرية التوزيع", "قسم العمليات", "شعبة الشمال",
             "EMP500", "محمد سعيد", "محاسب",
             "0799999999", "mohammed@company.com", "employee",
             "EMP010", "2026-01-01", "Yarmouk@2025"],
            ["مديرية التوزيع", "قسم العمليات", "شعبة الجنوب",
             "EMP501", "علي سعيد", "فني",
             "0799999998", "ali@company.com", "employee",
             "EMP011", "2026-01-15", "Yarmouk@2025"],
            ["مديرية الصيانة", "قسم الاصلاحات", "",
             "EMP502", "حسن محمود", "فني صيانة",
             "0799999997", "hassan@company.com", "employee",
             "EMP006", "2026-02-01", "Yarmouk@2025"],
        ]
        for row in sample_rows:
            ws.append(row)
        
        # Style header - green with white bold font
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Style sample rows
        sample_font = Font(size=11)
        sample_alignment = Alignment(horizontal="right", vertical="center")
        for row in ws.iter_rows(min_row=2, max_row=4):
            for cell in row:
                cell.font = sample_font
                cell.alignment = sample_alignment
                cell.border = thin_border
        
        # Auto-width columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 25)
        
        # Add data validation for role column
        role_validation = DataValidation(
            type="list",
            formula1='"employee,hr_manager,field_supervisor,office_supervisor,general_manager"',
            allow_blank=True
        )
        role_validation.error = "يرجى اختيار دور صحيح"
        role_validation.errorTitle = "خطأ في الدور"
        ws.add_data_validation(role_validation)
        role_validation.add(f"I2:I1000")
        
        # Add data validation for work type
        work_validation = DataValidation(
            type="list",
            formula1='"مكتب,ميدان,مختلط"',
            allow_blank=True
        )
        ws.add_data_validation(work_validation)
        
        # Add instructions sheet
        ws2 = wb.create_sheet(title="التعليمات")
        instructions = [
            ["تعليمات ملء القالب"],
            [""],
            ["=== التسلسل الإداري ==="],
            ["1. المديرية: اسم المديرية (مثال: مديرية التوزيع)"],
            ["2. القسم: اسم القسم (مثال: قسم العمليات)"],
            ["3. الشعبة: اسم الشعبة (مثال: شعبة الشمال) - اختياري"],
            [""],
            ["=== بيانات الموظف ==="],
            ["4. رقم الموظف: فريد ومقيد (مثال: EMP500)"],
            ["5. الاسم الكامل: بالعربي (مطلوب)"],
            ["6. المسمى الوظيفي: المسمى الوظيفي للموظف"],
            ["7. الهاتف: رقم الهاتف"],
            ["8. البريد الإلكتروني: البريد الإلكتروني"],
            ["9. الدور: employee / hr_manager / field_supervisor / office_supervisor / general_manager"],
            ["10. رقم المدير المباشر: رقم وظيفي المدير (مثال: EMP010)"],
            ["11. تاريخ التعيين: YYYY-MM-DD"],
            ["12. كلمة المرور: كلمة المرور للموظف (الافتراضي: Yarmouk@2025)"],
            [""],
            ["=== ملاحظات مهمة ==="],
            ["- الأعمدة المطلوبة: المديرية + القسم + رقم الموظف + الاسم الكامل + كلمة المرور"],
            ["- يتم تحديد الوحدة التنظيمية تلقائياً بناءً على التسلسل الإداري"],
            ["- يتم تخطي الموظفين المكررين تلقائياً"],
            ["- يتم إنشاء حساب للموظف تلقائياً"],
            ["- يمكنك استخدام القائمة المنسدلة للدور ونوع العمل"],
        ]
        for row in instructions:
            ws2.append(row)
        
        # Style instructions
        ws2["A1"].font = Font(bold=True, size=14, color="2E7D32")
        for row in ws2.iter_rows(min_row=2, max_row=len(instructions)):
            for cell in row:
                if cell.value:
                    cell.font = Font(size=11)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        from urllib.parse import quote
        encoded_filename = quote("قالب_الموظفين.xlsx")
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
    except ImportError:
        # CSV fallback
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["رقم الموظف", "الاسم الكامل", "المسمى الوظيفي", "الهاتف", "البريد الإلكتروني", "الدور", "الوحدة التنظيمية", "نوع العمل", "المدير المباشر", "تاريخ التعيين", "كلمة المرور"])
        writer.writerow(["EMP100", "أحمد محمد علي", "مهندس ميكانيكي", "0799999999", "ahmed@company.com", "employee", "الإدارة العامة", "مكتب", "EMP001", "2026-01-01", "Yarmouk@2025"])
        
        from urllib.parse import quote
        encoded_filename = quote("قالب_الموظفين.csv")
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )


@router.patch("/organization/units/{unit_id}")
async def update_org_unit(unit_id: int, body: dict, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(select(OrganizationUnit).where(OrganizationUnit.org_unit_id == unit_id))
    unit = result.scalar_one_or_none()
    if not unit:
        raise HTTPException(status_code=404, detail="Org unit not found")
    field_map = {"name": "unit_name", "unit_name": "unit_name", "unit_name_en": "unit_name_en", "parent_id": "parent_id"}
    for key, value in body.items():
        actual_key = field_map.get(key, key)
        if hasattr(unit, actual_key):
            setattr(unit, actual_key, value)
    await db.commit()
    return {"status": "updated", "org_unit_id": unit_id}


# ==================== REPORTS EXTENDED ====================

@router.get("/reports-extended/full-profile/{employee_id}")
async def full_profile(employee_id: int, format: str = "json", db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    result = await db.execute(select(EmployeeModel).where(EmployeeModel.employee_id == employee_id))
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"employee": {
        "employee_id": emp.employee_id, "employee_number": emp.employee_number,
        "full_name": emp.full_name, "job_title": emp.job_title,
    }}


@router.get("/reports-extended/attendance")
async def attendance_report(from_date: Optional[str] = None, to_date: Optional[str] = None, employee_id: Optional[int] = None, format: str = "json", db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    return {"items": [], "summary": {"total": 0}}


@router.get("/reports-extended/leave")
async def leave_report(from_date: Optional[str] = None, to_date: Optional[str] = None, status: Optional[str] = None, employee_id: Optional[int] = None, format: str = "json", db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    return {"items": [], "summary": {"total": 0}}


@router.get("/reports-extended/overtime")
async def overtime_report(from_date: Optional[str] = None, to_date: Optional[str] = None, status: Optional[str] = None, employee_id: Optional[int] = None, format: str = "json", db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    return {"items": [], "summary": {"total": 0}}


@router.get("/reports-extended/violations")
async def violations_report(from_date: Optional[str] = None, to_date: Optional[str] = None, penalty: Optional[str] = None, employee_id: Optional[int] = None, format: str = "json", db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    return {"items": [], "summary": {"total": 0}}


@router.get("/reports-extended/audit")
async def audit_report(from_date: Optional[str] = None, to_date: Optional[str] = None, employee_id: Optional[int] = None, action: Optional[str] = None, format: str = "json", db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    return {"items": [], "summary": {"total": 0}}


@router.get("/reports-extended/dashboard")
async def dashboard_report(org_unit_id: Optional[int] = None, from_date: Optional[str] = None, to_date: Optional[str] = None, db: AsyncSession = Depends(get_db), current_user: Employee = Depends(get_current_user)):
    from sqlalchemy import func
    total_emp = await db.execute(select(func.count(EmployeeModel.employee_id)))
    total = total_emp.scalar() or 0
    return {
        "total_employees": total,
        "present_today": 0, "on_leave": 0, "active_sessions": 0,
    }
