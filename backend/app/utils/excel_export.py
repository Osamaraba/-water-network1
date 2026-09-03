# -*- coding: utf-8 -*-
"""Excel export utilities for Yarmouk Water Management Pro."""
from datetime import datetime
from io import BytesIO
from typing import Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ARABIC_HEADERS = {
    # Generic
    "row": "م",
    "employee_number": "رقم الموظف",
    "full_name": "الاسم الكامل",
    "job_title": "المسمى الوظيفي",
    "org_unit": "الوحدة التنظيمية",
    "org_unit_name": "الوحدة التنظيمية",
    "work_type": "نوع العمل",
    "phone": "الهاتف",
    "email": "البريد الإلكتروني",
    "status": "الحالة",
    "is_active": "نشط",
    "hire_date": "تاريخ التعيين",
    # Attendance
    "check_in": "وقت الدخول",
    "check_out": "وقت الخروج",
    "check_in_time": "وقت الدخول",
    "check_out_time": "وقت الخروج",
    "duration_hours": "ساعات العمل",
    "work_duration_hours": "ساعات العمل",
    "attendance_status": "حالة الحضور",
    "present": "حاضر",
    "absent": "غائب",
    "on_leave": "في إجازة",
    "late": "متأخر",
    # Leave
    "leave_type": "نوع الإجازة",
    "leave_type_en": "نوع الإجازة (EN)",
    "leave_type_ar": "نوع الإجازة",
    "leave_type_custom": "تفاصيل النوع",
    "leave_kind": "نوع المغادرة",
    "leave_kind_en": "نوع المغادرة (EN)",
    "leave_kind_label": "نوع المغادرة",
    "outing_kind_en": "نوع المغادرة (EN)",
    "outing_kind_ar": "نوع المغادرة",
    "start_day": "يوم البداية",
    "end_day": "يوم النهاية",
    "outing_day": "يوم المغادرة",
    "start_date": "تاريخ البداية",
    "end_date": "تاريخ النهاية",
    "outing_date": "تاريخ المغادرة",
    "departure_time": "ساعة المغادرة",
    "return_time": "ساعة العودة",
    "leave_status": "حالة الإجازة",
    "start_time": "البدء",
    "end_time": "النهاية",
    "reason": "السبب",
    "review_note": "ملاحظة المراجع",
    "requested_days": "أيام مطلوبة",
    "approved_days": "أيام معتمدة",
    # Overtime
    "task_description": "وصف المهمة",
    "area_name": "اسم المنطقة",
    "requested_hours": "ساعات مطلوبة",
    "approved_hours": "ساعات معتمدة",
    "total_approved_hours": "ساعات معتمدة",
    "extended_hours": "ساعات ممتدة",
    "total_extended_hours": "ساعات ممتدة",
    "overtime_total_hours": "إجمالي الساعات (معتمدة+ممتد)",
    "overtime_status": "حالة العمل الإضافي",
    "tracking_starts_at": "بدء التتبع",
    "tracking_ends_at": "انتهاء التتبع",
    # Violations
    "violation_type": "نوع المخالفة",
    "violation_date": "تاريخ المخالفة",
    "violation_time": "وقت المخالفة",
    "penalty": "العقوبة",
    "penalty_label": "مستوى العقوبة",
    "notes": "ملاحظات",
    # GPS / Tracking
    "session_id": "معرّف الجلسة",
    "tracking_type": "نوع التتبع",
    "track_mode": "وضع التتبع",
    "track_interval": "الفاصل",
    "started_at": "بدء الجلسة",
    "ended_at": "انتهاء الجلسة",
    "is_outside": "خارج النطاق",
    "outside_distance_m": "مسافة الخروج (م)",
    "outside_started_at": "وقت بدء الخروج",
    "latest_lat": "آخر خط عرض",
    "latest_lng": "آخر خط طول",
    "breach_count": "عدد الخروقات",
    "total_outside_seconds": "إجمالي وقت الخروج (ث)",
    "total_outside_distance_m": "إجمالي مسافة الخروج (م)",
    "destination": "الوجهة",
    "tracking_required": "تتبع GPS",
    "tracking_acknowledged": "تم تأكيد التتبع",
    # Audit / activity
    "action": "الإجراء",
    "entity_type": "الكيان",
    "entity_id": "معرّف الكيان",
    "ip_address": "عنوان IP",
    "user_agent": "وكيل المستخدم",
    "created_at": "تاريخ الإنشاء",
    "report_date": "تاريخ التقرير",
    # Full profile / summary
    "field": "الحقل",
    "value": "القيمة",
    "metric": "المقياس",
    "section": "القسم",
    "description": "الوصف",
    "day": "اليوم",
    "in": "دخول",
    "out": "خروج",
    "summary": "الملخّص",
    "count": "العدد",
    # Common
    "name": "الاسم",
    "username": "اسم المستخدم",
    "role": "الدور",
    "type": "النوع",
}


def _header_fill():
    return PatternFill(start_color="1E4D8C", end_color="1E4D8C", fill_type="solid")


def _alt_fill():
    return PatternFill(start_color="F1F5FA", end_color="F1F5FA", fill_type="solid")


def _warning_fill():
    return PatternFill(start_color="FFE9B0", end_color="FFE9B0", fill_type="solid")


def _success_fill():
    return PatternFill(start_color="D6F0E0", end_color="D6F0E0", fill_type="solid")


def _danger_fill():
    return PatternFill(start_color="F8D2D2", end_color="F8D2D2", fill_type="solid")


def _thin_border():
    side = Side(border_style="thin", color="BFC8D4")
    return Border(left=side, right=side, top=side, bottom=side)


def _write_title(ws, title: str, last_col_letter: str) -> None:
    ws.merge_cells(f"A1:{last_col_letter}1")
    cell = ws["A1"]
    cell.value = title
    cell.font = Font(name="Arial", size=16, bold=True, color="1E4D8C")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28


def _write_subtitle(ws, subtitle: str, last_col_letter: str) -> None:
    ws.merge_cells(f"A2:{last_col_letter}2")
    cell = ws["A2"]
    cell.value = subtitle
    cell.font = Font(name="Arial", size=11, italic=True, color="5C6B82")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20


def _translate_headers(headers: Iterable[str]) -> List[str]:
    return [ARABIC_HEADERS.get(h, h) for h in headers]


def _autosize(ws, max_width: int = 42) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            # crude width: count visible chars, treat Arabic as wider
            ar_chars = sum(1 for c in s if "\u0600" <= c <= "\u06FF")
            en_chars = len(s) - ar_chars
            estimate = ar_chars * 1.6 + en_chars * 1.0
            if estimate > max_len:
                max_len = estimate
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), max_width)


def _write_header_row(ws, headers: List[str], row: int = 3) -> int:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = _header_fill()
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[row].height = 24
    return row + 1


def _write_data_rows(ws, rows: List[List], start_row: int) -> int:
    for r_offset, row_values in enumerate(rows):
        excel_row = start_row + r_offset
        fill = _alt_fill() if r_offset % 2 == 1 else None
        for c_offset, value in enumerate(row_values, start=1):
            cell = ws.cell(row=excel_row, column=c_offset, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _thin_border()
            if fill:
                cell.fill = fill
    return start_row + len(rows)


def _filename(prefix: str, ext: str = "xlsx") -> str:
    return f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"


def build_workbook_from_table(
    title: str,
    subtitle: str,
    headers_en: List[str],
    rows: List[List],
    filename_prefix: str = "report",
) -> Tuple[bytes, str]:
    """Build a single-sheet XLSX from headers + rows. Returns (bytes, filename)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = _translate_headers(headers_en)
    last_col_letter = get_column_letter(max(1, len(headers)))

    _write_title(ws, title, last_col_letter)
    _write_subtitle(ws, subtitle, last_col_letter)
    next_row = _write_header_row(ws, headers, row=3)
    _write_data_rows(ws, rows, next_row)
    _autosize(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), _filename(filename_prefix)


def build_multi_sheet_workbook(
    sections: List[dict],
    filename_prefix: str = "report",
) -> Tuple[bytes, str]:
    """Build a multi-sheet XLSX. sections = [{title, subtitle, headers_en, rows, sheet_name?}]"""
    wb = Workbook()
    # remove default
    wb.remove(wb.active)
    for idx, sec in enumerate(sections):
        ws_name = sec.get("sheet_name") or f"Sheet{idx+1}"
        # Excel sheet names max 31 chars
        ws_name = ws_name[:31]
        ws = wb.create_sheet(ws_name)
        headers = _translate_headers(sec["headers_en"])
        last_col_letter = get_column_letter(max(1, len(headers)))
        _write_title(ws, sec.get("title", "Report"), last_col_letter)
        _write_subtitle(ws, sec.get("subtitle", ""), last_col_letter)
        next_row = _write_header_row(ws, headers, row=3)
        _write_data_rows(ws, sec.get("rows", []), next_row)
        _autosize(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), _filename(filename_prefix)
