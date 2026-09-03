# -*- coding: utf-8 -*-
"""Generate 4 Excel files for Irbid, Ajloun, Jerash, Mafraq.

Each row uses Arabic-friendly columns:
  - org_unit_name (Arabic full text, e.g. "مديرية توزيع المياه - إربد")
  - job_title (short title, e.g. "مدير")
  - The Backend will auto-derive codes via OrgUnitResolver + JobTitleResolver.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("A", "employee_number",  "رقم الموظف *",            16),
    ("B", "full_name",        "الاسم الكامل *",          28),
    ("C", "full_name_en",     "Name (EN)",               22),
    ("D", "job_title",        "المسمى الوظيفي *",         24),
    ("E", "phone",            "الهاتف",                 14),
    ("F", "email",            "البريد",                 26),
    ("G", "hire_date",        "تاريخ التعيين",          14),
    ("H", "role",             "الدور *",                 22),
    ("I", "org_unit_name",    "الوحدة التنظيمية *",     28),
    ("J", "org_unit_type",    "نوع الوحدة *",          16),
    ("K", "direct_manager_num","رقم المدير *",           20),
    ("L", "work_type",        "نوع العمل *",            14),
    ("M", "allow_tracking",   "GPS",                    16),
    ("N", "geofence_lat",     "Latitude",               14),
    ("O", "geofence_lng",     "Longitude",              14),
    ("P", "geofence_radius",  "Radius (m)",             14),
    ("Q", "geofence_exempt",  "Exempt",                 16),
    ("R", "password",         "كلمة المرور *",          20),
    ("S", "notes",            "ملاحظات",                28),
]

ROLE_LIST = "general_manager,hr_manager,field_supervisor,office_supervisor,employee"
UNIT_LIST  = "COMPANY,PROVINCE,DIRECTORATE,SECTION,SUBSECTION,UNIT"
WORK_LIST  = "FIELD,OFFICE,HYBRID"
BOOL_LIST  = "TRUE,FALSE"


def fill(c): return PatternFill(start_color=c, end_color=c, fill_type="solid")
def bdr():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def build_province_file(province_ar, province_en, emp_prefix, employees, output_dir):
    wb = Workbook()
    wb.remove(wb.active)

    # === Sheet 1: Instructions ===
    ws1 = wb.create_sheet("1-Instructions")
    ws1.sheet_view.showGridLines = False
    for i, w in enumerate([5, 60, 30], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.row_dimensions[1].height = 40
    ws1.merge_cells("A1:C1")
    c = ws1["A1"]
    c.value = f"محافظة {province_ar} — إدخال الموظفين"
    c.font = Font(name="Arial", size=18, bold=True, color="FFFFFF")
    c.fill = fill("1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")

    info = [
        (f"عدد الموظفين في هذا الملف: {len(employees)}", "2E75B6", True),
        ("كل ملف محافظة يُستورد عبر /employees/bulk-import", None),
        ("النظام يستخرج كود الوحدة (DIR-IRB-WAT) وكود المسمى (WAT-MGR) تلقائياً", None),
        ("", None),
        ("الهيكل:", "2E75B6", True),
        ("1. مدير المحافظة (PROVINCE) — كود: PROV-IRB", None),
        ("2. مدير مديرية توزيع المياه (DIRECTORATE) — كود: DIR-IRB-WAT", None),
        ("3. مدير مديرية الصرف الصحي (DIRECTORATE) — كود: DIR-IRB-SAN", None),
        ("4. مدير خدمات المشتركين (DIRECTORATE) — كود: DIR-IRB-CUS", None),
        ("5. الصيانة (UNIT) — كود: UNT-IRB-MNT-01-1", None),
    ]
    for i, item in enumerate(info, 3):
        text, bg = item[0], item[1]
        bold = item[2] if len(item) > 2 else False
        ws1.row_dimensions[i].height = 22
        ws1.merge_cells(f"A{i}:C{i}")
        c = ws1[f"A{i}"]
        c.value = text
        if bg and bold:
            c.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            c.fill = fill(bg)
        else:
            c.font = Font(name="Arial", size=10, color="404040")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # === Sheet 2: DataEntry ===
    ws2 = wb.create_sheet("DataEntry")
    ws2.sheet_view.showGridLines = False
    num_cols = len(COLUMNS)
    last_col = get_column_letter(num_cols)

    ws2.row_dimensions[1].height = 40
    ws2.merge_cells(f"A1:{last_col}1")
    c = ws2["A1"]
    c.value = f"موظفو محافظة {province_ar} — البيانات مُعبّأة (الكود يُولّد تلقائياً)"
    c.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    c.fill = fill("1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws2.row_dimensions[2].height = 36
    for col_idx, (col_letter, _, header, width) in enumerate(COLUMNS, 1):
        required = header.endswith(" *")
        bg = "1F3864" if required else "2E75B6"
        c = ws2.cell(row=2, column=col_idx, value=header)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr()
        ws2.column_dimensions[col_letter].width = width

    start_row = 3
    for i, emp in enumerate(employees):
        row = start_row + i
        ws2.row_dimensions[row].height = 22
        for col_idx, val in enumerate(emp, 1):
            c = ws2.cell(row=row, column=col_idx, value=val)
            c.border = bdr()
            c.alignment = Alignment(horizontal="center", vertical="center")
            if i % 2 == 1:
                c.fill = fill("F7FBFF")

    buf_path = os.path.join(output_dir, f"province_{province_en}_{emp_prefix}.xlsx")
    wb.save(buf_path)
    return buf_path


# ============================================================================
# IRBID - 12 employees
# ============================================================================
IRBID = [
    # 1. Province Director
    [f"IRB101", "محمد إربد", "Mohammed Irbid", "مدير", "0791000101",
     "m.irbid@yarmouk.gov.jo", "2020-06-15", "hr_manager",
     "محافظة إربد", "PROVINCE", "EMP001", "OFFICE", "FALSE", "", "", "200", "FALSE",
     "Yarmouk@2025", "مدير إربد"],

    # 2-5. Water Distribution Directorate
    [f"IRB102", "أحمد التوزيع", "Ahmed Tazeez", "مدير", "0791000102",
     "a.tazeez@yarmouk.gov.jo", "2021-02-01", "field_supervisor",
     "مديرية توزيع المياه - إربد", "DIRECTORATE", "IRB101", "FIELD", "TRUE",
     "32.5550", "35.8500", "200", "FALSE", "Yarmouk@2025", "مدير توزيع"],
    [f"IRB103", "فاطمة التوزيع", "Fatima Tazeez", "رئيس قسم", "0791000103",
     "f.tazeez@yarmouk.gov.jo", "2021-05-20", "office_supervisor",
     "قسم توزيع 1 - إربد", "SECTION", "IRB102", "FIELD", "TRUE",
     "32.5560", "35.8510", "150", "FALSE", "Yarmouk@2025", "رئيس قسم"],
    [f"IRB104", "علي الفني", "Ali Tech", "فني", "0791000104",
     "a.tech@yarmouk.gov.jo", "2022-01-10", "employee",
     "وحدة توزيع 1 - إربد", "UNIT", "IRB103", "FIELD", "TRUE",
     "32.5570", "35.8520", "100", "FALSE", "Yarmouk@2025", "فني ميداني"],
    [f"IRB105", "نور الفني", "Noor Tech", "فني", "0791000105",
     "n.tech@yarmouk.gov.jo", "2022-03-15", "employee",
     "وحدة توزيع 1 - إربد", "UNIT", "IRB103", "FIELD", "TRUE",
     "32.5580", "35.8530", "100", "FALSE", "Yarmouk@2025", "فني ميداني"],

    # 6-8. Sanitation Directorate
    [f"IRB106", "سعد الصرف", "Saad Seref", "مدير", "0791000106",
     "s.seref@yarmouk.gov.jo", "2021-03-01", "field_supervisor",
     "مديرية الصرف الصحي - إربد", "DIRECTORATE", "IRB101", "FIELD", "TRUE",
     "32.5550", "35.8500", "200", "FALSE", "Yarmouk@2025", "مدير صرف"],
    [f"IRB107", "مريم الصرف", "Maryam Seref", "رئيس قسم", "0791000107",
     "m.seref@yarmouk.gov.jo", "2021-07-20", "office_supervisor",
     "قسم صرف 1 - إربد", "SECTION", "IRB106", "FIELD", "TRUE",
     "32.5560", "35.8510", "150", "FALSE", "Yarmouk@2025", "رئيس قسم"],
    [f"IRB108", "خالد الصرف", "Khaled Seref", "فني", "0791000108",
     "k.seref@yarmouk.gov.jo", "2022-02-10", "employee",
     "وحدة صرف 1 - إربد", "UNIT", "IRB107", "FIELD", "TRUE",
     "32.5570", "35.8520", "100", "FALSE", "Yarmouk@2025", "فني صرف"],

    # 9-11. Customer Services
    [f"IRB109", "حسن المشتركين", "Hassan Moshtareen", "مدير", "0791000109",
     "h.mosh@yarmouk.gov.jo", "2021-04-01", "field_supervisor",
     "مديرية خدمات المشتركين - إربد", "DIRECTORATE", "IRB101", "OFFICE", "FALSE",
     "32.5550", "35.8500", "200", "FALSE", "Yarmouk@2025", "مدير خدمات"],
    [f"IRB110", "ليلى الخدمات", "Layla Services", "موظف", "0791000110",
     "l.serv@yarmouk.gov.jo", "2022-05-15", "employee",
     "وحدة خدمات المشتركين - إربد", "UNIT", "IRB109", "OFFICE", "FALSE",
     "", "", "200", "FALSE", "Yarmouk@2025", "موظفة خدمات"],
    [f"IRB111", "يوسف الخدمات", "Youssef Services", "موظف", "0791000111",
     "y.serv@yarmouk.gov.jo", "2022-06-15", "employee",
     "وحدة خدمات المشتركين - إربد", "UNIT", "IRB109", "OFFICE", "FALSE",
     "", "", "200", "FALSE", "Yarmouk@2025", "موظف خدمات"],

    # 12. Maintenance
    [f"IRB112", "رامي الصيانة", "Rami Siana", "فني", "0791000112",
     "r.siana@yarmouk.gov.jo", "2022-08-01", "employee",
     "وحدة صيانة - إربد", "UNIT", "IRB101", "FIELD", "TRUE",
     "32.5550", "35.8500", "200", "FALSE", "Yarmouk@2025", "فني صيانة"],
]

# ============================================================================
# AJLOUN - 10 employees
# ============================================================================
AJLOUN = [
    ["AJL101", "كريم عجلون", "Kareem Ajloun", "مدير", "0791000201",
     "k.ajloun@yarmouk.gov.jo", "2020-07-01", "hr_manager",
     "محافظة عجلون", "PROVINCE", "EMP001", "OFFICE", "FALSE", "", "", "200", "FALSE",
     "Yarmouk@2025", "مدير عجلون"],

    ["AJL102", "هاني التوزيع", "Hani Tazeez", "مدير", "0791000202",
     "h.tazeez@yarmouk.gov.jo", "2021-03-15", "field_supervisor",
     "مديرية توزيع المياه - عجلون", "DIRECTORATE", "AJL101", "FIELD", "TRUE",
     "32.3325", "35.7512", "200", "FALSE", "Yarmouk@2025", "مدير توزيع"],
    ["AJL103", "سمير التوزيع", "Sameer Tazeez", "رئيس قسم", "0791000203",
     "s.tazeez@yarmouk.gov.jo", "2021-08-01", "office_supervisor",
     "قسم توزيع 1 - عجلون", "SECTION", "AJL102", "FIELD", "TRUE",
     "32.3325", "35.7512", "150", "FALSE", "Yarmouk@2025", "رئيس قسم"],
    ["AJL104", "وسيم الفني", "Waseem Tech", "فني", "0791000204",
     "w.tech@yarmouk.gov.jo", "2022-04-10", "employee",
     "وحدة توزيع 1 - عجلون", "UNIT", "AJL103", "FIELD", "TRUE",
     "32.3325", "35.7512", "100", "FALSE", "Yarmouk@2025", "فني ميداني"],

    ["AJL105", "نادر الصرف", "Nader Seref", "مدير", "0791000205",
     "n.seref@yarmouk.gov.jo", "2021-04-01", "field_supervisor",
     "مديرية الصرف الصحي - عجلون", "DIRECTORATE", "AJL101", "FIELD", "TRUE",
     "32.3325", "35.7512", "200", "FALSE", "Yarmouk@2025", "مدير صرف"],
    ["AJL106", "ريما الصرف", "Reema Seref", "رئيس قسم", "0791000206",
     "r.seref@yarmouk.gov.jo", "2021-09-01", "office_supervisor",
     "قسم صرف 1 - عجلون", "SECTION", "AJL105", "FIELD", "TRUE",
     "32.3325", "35.7512", "150", "FALSE", "Yarmouk@2025", "رئيسة قسم"],
    ["AJL107", "بسام الصرف", "Bassam Seref", "فني", "0791000207",
     "b.seref@yarmouk.gov.jo", "2022-05-15", "employee",
     "وحدة صرف 1 - عجلون", "UNIT", "AJL106", "FIELD", "TRUE",
     "32.3325", "35.7512", "100", "FALSE", "Yarmouk@2025", "فني صرف"],

    ["AJL108", "لينا المشتركين", "Lina Moshtareen", "مدير", "0791000208",
     "l.ajloun@yarmouk.gov.jo", "2021-05-15", "field_supervisor",
     "مديرية خدمات المشتركين - عجلون", "DIRECTORATE", "AJL101", "OFFICE", "FALSE",
     "32.3325", "35.7512", "200", "FALSE", "Yarmouk@2025", "مديرة خدمات"],
    ["AJL109", "مجد الخدمات", "Majd Services", "موظف", "0791000209",
     "m.ajloun@yarmouk.gov.jo", "2022-07-01", "employee",
     "وحدة خدمات المشتركين - عجلون", "UNIT", "AJL108", "OFFICE", "FALSE",
     "", "", "200", "FALSE", "Yarmouk@2025", "موظف خدمات"],

    ["AJL110", "عصام الصيانة", "Essam Siana", "فني", "0791000210",
     "e.siana@yarmouk.gov.jo", "2022-09-01", "employee",
     "وحدة صيانة - عجلون", "UNIT", "AJL101", "FIELD", "TRUE",
     "32.3325", "35.7512", "200", "FALSE", "Yarmouk@2025", "فني صيانة"],
]

# ============================================================================
# JERASH - 10 employees
# ============================================================================
JERASH = [
    ["JER101", "عصام جرش", "Essam Jerash", "مدير", "0791000301",
     "e.jerash@yarmouk.gov.jo", "2020-08-01", "hr_manager",
     "محافظة جرش", "PROVINCE", "EMP001", "OFFICE", "FALSE", "", "", "200", "FALSE",
     "Yarmouk@2025", "مدير جرش"],

    ["JER102", "تيسير التوزيع", "Tayseer Tazeez", "مدير", "0791000302",
     "t.tazeez@yarmouk.gov.jo", "2021-04-01", "field_supervisor",
     "مديرية توزيع المياه - جرش", "DIRECTORATE", "JER101", "FIELD", "TRUE",
     "32.2806", "35.8993", "200", "FALSE", "Yarmouk@2025", "مدير توزيع"],
    ["JER103", "ميساء التوزيع", "Maysaa Tazeez", "رئيس قسم", "0791000303",
     "m.tazeez@yarmouk.gov.jo", "2021-10-01", "office_supervisor",
     "قسم توزيع 1 - جرش", "SECTION", "JER102", "FIELD", "TRUE",
     "32.2806", "35.8993", "150", "FALSE", "Yarmouk@2025", "رئيسة قسم"],
    ["JER104", "عدنان الفني", "Adnan Tech", "فني", "0791000304",
     "a.tech@yarmouk.gov.jo", "2022-05-15", "employee",
     "وحدة توزيع 1 - جرش", "UNIT", "JER103", "FIELD", "TRUE",
     "32.2806", "35.8993", "100", "FALSE", "Yarmouk@2025", "فني ميداني"],

    ["JER105", "نضال الصرف", "Nidal Seref", "مدير", "0791000305",
     "n.seref@yarmouk.gov.jo", "2021-05-01", "field_supervisor",
     "مديرية الصرف الصحي - جرش", "DIRECTORATE", "JER101", "FIELD", "TRUE",
     "32.2806", "35.8993", "200", "FALSE", "Yarmouk@2025", "مدير صرف"],
    ["JER106", "إيمان الصرف", "Eman Seref", "رئيس قسم", "0791000306",
     "i.seref@yarmouk.gov.jo", "2021-11-01", "office_supervisor",
     "قسم صرف 1 - جرش", "SECTION", "JER105", "FIELD", "TRUE",
     "32.2806", "35.8993", "150", "FALSE", "Yarmouk@2025", "رئيسة قسم"],
    ["JER107", "فادي الصرف", "Fadi Seref", "فني", "0791000307",
     "f.seref@yarmouk.gov.jo", "2022-06-15", "employee",
     "وحدة صرف 1 - جرش", "UNIT", "JER106", "FIELD", "TRUE",
     "32.2806", "35.8993", "100", "FALSE", "Yarmouk@2025", "فني صرف"],

    ["JER108", "رانية المشتركين", "Rania Moshtareen", "مدير", "0791000308",
     "r.jerash@yarmouk.gov.jo", "2021-06-15", "field_supervisor",
     "مديرية خدمات المشتركين - جرش", "DIRECTORATE", "JER101", "OFFICE", "FALSE",
     "32.2806", "35.8993", "200", "FALSE", "Yarmouk@2025", "مديرة خدمات"],
    ["JER109", "باسم الخدمات", "Basem Services", "موظف", "0791000309",
     "b.jerash@yarmouk.gov.jo", "2022-08-01", "employee",
     "وحدة خدمات المشتركين - جرش", "UNIT", "JER108", "OFFICE", "FALSE",
     "", "", "200", "FALSE", "Yarmouk@2025", "موظف خدمات"],

    ["JER110", "رائد الصيانة", "Raed Siana", "فني", "0791000310",
     "r.siana@yarmouk.gov.jo", "2022-10-01", "employee",
     "وحدة صيانة - جرش", "UNIT", "JER101", "FIELD", "TRUE",
     "32.2806", "35.8993", "200", "FALSE", "Yarmouk@2025", "فني صيانة"],
]

# ============================================================================
# MAFRAQ - 12 employees
# ============================================================================
MAFRAQ = [
    ["MAF101", "إبراهيم المفرق", "Ibrahim Mafraq", "مدير", "0791000401",
     "i.mafraq@yarmouk.gov.jo", "2020-09-01", "hr_manager",
     "محافظة المفرق", "PROVINCE", "EMP001", "OFFICE", "FALSE", "", "", "200", "FALSE",
     "Yarmouk@2025", "مدير المفرق"],

    ["MAF102", "سامي التوزيع", "Sami Tazeez", "مدير", "0791000402",
     "s.tazeez@yarmouk.gov.jo", "2021-05-01", "field_supervisor",
     "مديرية توزيع المياه - المفرق", "DIRECTORATE", "MAF101", "FIELD", "TRUE",
     "32.3422", "36.2080", "250", "FALSE", "Yarmouk@2025", "مدير توزيع"],
    ["MAF103", "نواف التوزيع", "Nawaf Tazeez", "رئيس قسم", "0791000403",
     "n.tazeez@yarmouk.gov.jo", "2021-12-01", "office_supervisor",
     "قسم توزيع 1 - المفرق", "SECTION", "MAF102", "FIELD", "TRUE",
     "32.3422", "36.2080", "150", "FALSE", "Yarmouk@2025", "رئيس قسم"],
    ["MAF104", "راشد الفني", "Rashed Tech", "فني", "0791000404",
     "r.tech@yarmouk.gov.jo", "2022-06-15", "employee",
     "وحدة توزيع 1 - المفرق", "UNIT", "MAF103", "FIELD", "TRUE",
     "32.3422", "36.2080", "100", "FALSE", "Yarmouk@2025", "فني ميداني"],
    ["MAF105", "ماجد الفني", "Majed Tech", "فني", "0791000405",
     "m.tech@yarmouk.gov.jo", "2022-07-15", "employee",
     "وحدة توزيع 1 - المفرق", "UNIT", "MAF103", "FIELD", "TRUE",
     "32.3422", "36.2080", "100", "FALSE", "Yarmouk@2025", "فني ميداني"],

    ["MAF106", "عبدالله الصرف", "Abdullah Seref", "مدير", "0791000406",
     "a.seref@yarmouk.gov.jo", "2021-06-01", "field_supervisor",
     "مديرية الصرف الصحي - المفرق", "DIRECTORATE", "MAF101", "FIELD", "TRUE",
     "32.3422", "36.2080", "250", "FALSE", "Yarmouk@2025", "مدير صرف"],
    ["MAF107", "حمد الصرف", "Hamad Seref", "رئيس قسم", "0791000407",
     "h.seref@yarmouk.gov.jo", "2022-01-01", "office_supervisor",
     "قسم صرف 1 - المفرق", "SECTION", "MAF106", "FIELD", "TRUE",
     "32.3422", "36.2080", "150", "FALSE", "Yarmouk@2025", "رئيس قسم"],
    ["MAF108", "وليد الصرف", "Walid Seref", "فني", "0791000408",
     "w.seref@yarmouk.gov.jo", "2022-08-15", "employee",
     "وحدة صرف 1 - المفرق", "UNIT", "MAF107", "FIELD", "TRUE",
     "32.3422", "36.2080", "100", "FALSE", "Yarmouk@2025", "فني صرف"],

    ["MAF109", "فهد المشتركين", "Fahad Moshtareen", "مدير", "0791000409",
     "f.mafraq@yarmouk.gov.jo", "2021-07-15", "field_supervisor",
     "مديرية خدمات المشتركين - المفرق", "DIRECTORATE", "MAF101", "OFFICE", "FALSE",
     "32.3422", "36.2080", "250", "FALSE", "Yarmouk@2025", "مدير خدمات"],
    ["MAF110", "ريم الخدمات", "Reem Services", "موظف", "0791000410",
     "r.mafraq@yarmouk.gov.jo", "2022-09-01", "employee",
     "وحدة خدمات المشتركين - المفرق", "UNIT", "MAF109", "OFFICE", "FALSE",
     "", "", "200", "FALSE", "Yarmouk@2025", "موظفة خدمات"],
    ["MAF111", "خالد الخدمات", "Khaled Services", "موظف", "0791000411",
     "k.mafraq@yarmouk.gov.jo", "2022-10-01", "employee",
     "وحدة خدمات المشتركين - المفرق", "UNIT", "MAF109", "OFFICE", "FALSE",
     "", "", "200", "FALSE", "Yarmouk@2025", "موظف خدمات"],

    ["MAF112", "منصور الصيانة", "Mansour Siana", "فني", "0791000412",
     "m.siana@yarmouk.gov.jo", "2022-11-01", "employee",
     "وحدة صيانة - المفرق", "UNIT", "MAF101", "FIELD", "TRUE",
     "32.3422", "36.2080", "250", "FALSE", "Yarmouk@2025", "فني صيانة"],
]


if __name__ == "__main__":
    output_dir = r"D:\yarmouk_water_management_pro\backend\output\provinces"
    os.makedirs(output_dir, exist_ok=True)

    files = []
    files.append(build_province_file("إربد", "irbid", "IRB", IRBID, output_dir))
    files.append(build_province_file("عجلون", "ajloun", "AJL", AJLOUN, output_dir))
    files.append(build_province_file("جرش", "jerash", "JER", JERASH, output_dir))
    files.append(build_province_file("المفرق", "mafraq", "MAF", MAFRAQ, output_dir))

    print("Generated files:")
    for f in files:
        size_kb = os.path.getsize(f) / 1024
        print(f"  {os.path.basename(f):40s}  {size_kb:.1f} KB")
    total = sum([len(IRBID), len(AJLOUN), len(JERASH), len(MAFRAQ)])
    print(f"\nTotal: {len(files)} files, {total} employees")
