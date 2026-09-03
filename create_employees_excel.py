import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "الموظفين"

headers = [
    "الرقم الوظيفي", "الاسم الكامل", "الاسم بالإنجليزية", "رقم الهاتف",
    "المسمى الوظيفي", "الوحدة التنظيمية (كود)", "نوع العمل",
    "المدير المباشر (رقم وظيفي)", "البريد الإلكتروني"
]

header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="1E4D8C", end_color="1E4D8C", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# 30 employees with diverse job titles
# org codes: GM, HR, DIST, SEW, CUST, MAINT, DIST-SEC1, SEW-SEC1, CUST-SEC1, MAINT-SEC1
employees = [
    ("GM001", "عبدالله الخطيب", "Abdullah Al-Khatib", "0799001001", "المدير العام", "GM", "مكتبي", "GM001", "gm@yarmouk.jo"),
    ("HR001", "نورة الحوراني", "Noura Al-Hourani", "0799001002", "مدير الموارد البشرية", "HR", "مكتبي", "GM001", "hr@yarmouk.jo"),
    ("HR002", "سحر البياتي", "Sahar Al-Bayati", "0799001003", "pegawai موارد بشرية", "HR", "مكتبي", "HR001", "hr2@yarmouk.jo"),
    ("DIR001", "خالد العمري", "Khaled Al-Omari", "0799001004", "مدير توزيع المياه", "DIST", "مكتبي", "GM001", "dist@yarmouk.jo"),
    ("DIR002", "ياسمين الخطيب", "Yasmine Al-Khatib", "0799001005", "مدير الصرف الصحي", "SEW", "مكتبي", "GM001", "sew@yarmouk.jo"),
    ("DIR003", "مصطفى الزعبي", "Mustafa Al-Zoubi", "0799001006", "مدير خدمات الزبائن", "CUST", "مكتبي", "GM001", "cust@yarmouk.jo"),
    ("DIR004", "أحمد الصمادي", "Ahmad Al-Samadi", "0799001007", "مدير الصيانة والتوزيع", "MAINT", "مكتبي", "GM001", "maint@yarmouk.jo"),
    ("SEC001", "رائد الحمود", "Raid Al-Hamoud", "0799001008", "رئيس قسم توزيع المياه", "DIST-SEC1", "ميداني", "DIR001", "s1@yarmouk.jo"),
    ("SEC002", "أمل الدوسري", "Amal Al-Dosari", "0799001009", "رئيس قسم الصرف الصحي", "SEW-SEC1", "ميداني", "DIR002", "s2@yarmouk.jo"),
    ("SEC003", "عمران الشخانة", "Imran Al-Shakhanah", "0799001010", "رئيس قسم خدمات الزبائن", "CUST-SEC1", "مكتبي", "DIR003", "s3@yarmouk.jo"),
    ("SEC004", "حمزة القضاة", "Hamza Al-Qudah", "0799001011", "رئيس قسم الصيانة", "MAINT-SEC1", "ميداني", "DIR004", "s4@yarmouk.jo"),
    ("DWN001", "ديمة الحلبي", "Dima Al-Halabi", "0799001012", "كاتب ديوان", "GM", "مكتبي", "GM001", "dwn@yarmouk.jo"),
    ("WRK001", "عثمان الريالات", "Othman Al-Ryalat", "0799001013", "حداد مسرجي", "MAINT-SEC1", "ميداني", "SEC004", "w1@yarmouk.jo"),
    ("WRK002", "طارق المعاني", "Tariq Al-Maaytah", "0799001014", "عامل صيانة", "MAINT-SEC1", "ميداني", "SEC004", "w2@yarmouk.jo"),
    ("MTR001", "ندى الخطيب", "Nada Al-Khatib", "0799001015", "قارئ عدادات", "DIST-SEC1", "ميداني", "SEC001", "m1@yarmouk.jo"),
    ("MTR002", "وسيم المقيت", "Waseem Al-Maqt", "0799001016", "قارئ عدادات", "DIST-SEC1", "ميداني", "SEC001", "m2@yarmouk.jo"),
    ("COL001", "ريم عبود", "Reem Abboud", "0799001017", "محصل", "CUST-SEC1", "ميداني", "SEC003", "c1@yarmouk.jo"),
    ("COL002", "فادي النجار", "Fadi Al-Najjar", "0799001018", "محصل", "CUST-SEC1", "ميداني", "SEC003", "c2@yarmouk.jo"),
    ("WAJ001", "هيثم الطراونة", "Haitham Al-Taraawneh", "0799001019", "وايجي", "CUST-SEC1", "ميداني", "SEC003", "wj1@yarmouk.jo"),
    ("DRV001", "مصطفى المشهور", "Mostafa Al-Mashhour", "0799001020", "سائق", "DIST", "ميداني", "DIR001", "dr1@yarmouk.jo"),
    ("OPR001", "نبيل الجيوسي", "Nabil Al-Jayousi", "0799001021", "مشغل محطة", "DIST-SEC1", "ميداني", "SEC001", "op1@yarmouk.jo"),
    ("WAT001", "سالم اليدوبين", "Salem Al-Yadoubin", "0799001022", "موزع مياه", "DIST-SEC1", "ميداني", "SEC001", "wt1@yarmouk.jo"),
    ("WAT002", "منى العتوم", "Mona Al-Atoum", "0799001023", "موزعة مياه", "DIST-SEC1", "ميداني", "SEC001", "wt2@yarmouk.jo"),
    ("FRG001", "حسام بدارنة", "Hussam Badarneh", "0799001024", "فني توزيع مياه", "DIST-SEC1", "ميداني", "SEC001", "fg1@yarmouk.jo"),
    ("FRG002", "لutfi العمري", "Lutfi Al-Aamri", "0799001025", "فني صرف صحي", "SEW-SEC1", "ميداني", "SEC002", "fg2@yarmouk.jo"),
    ("SUP001", "يحيى الغزاوي", "Yahya Al-Ghazawi", "0799001026", "مشرف ميداني", "DIST-SEC1", "ميداني", "SEC001", "sp1@yarmouk.jo"),
    ("SUP002", "محمد الخزاعلة", "Muhannad Al-Khazaleh", "0799001027", "مشرف ميداني", "SEW-SEC1", "ميداني", "SEC002", "sp2@yarmouk.jo"),
    ("ACC001", "منال الشهاب", "Manal Al-Shahab", "0799001028", "محاسبة", "GM", "مكتبي", "HR001", "ac1@yarmouk.jo"),
    ("DWN002", "عيسى المهتدي", "Issa Al-Mohtadi", "0799001029", "رئيس ديوان", "GM", "مكتبي", "GM001", "dw2@yarmouk.jo"),
    ("ADM001", "ريم العزام", "Reem Al-Ezzam", "0799001030", "سكرتيرة", "HR", "مكتبي", "HR001", "ad1@yarmouk.jo"),
]

for row_idx, emp in enumerate(employees, 2):
    for col_idx, val in enumerate(emp, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 30
ws.column_dimensions['F'].width = 20
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 20
ws.column_dimensions['I'].width = 25

output_path = r"D:\yarmouk_water_management_pro\employees_import.xlsx"
wb.save(output_path)
print("Excel file saved:", output_path)
print("Total employees:", len(employees))
